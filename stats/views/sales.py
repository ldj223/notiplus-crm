import os, json
from datetime import datetime, timedelta, date
from django.conf import settings
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count
from django.contrib import messages
from django.utils import timezone
import openpyxl
import xlrd
import calendar
from django.core.exceptions import ValidationError
from decimal import Decimal, ROUND_HALF_UP
import logging
from rest_framework.decorators import api_view
from django.db.models import Q
from django.core.cache import cache

from ..forms import CredentialForm, SignUpForm
from ..models import (
    AdStats, PlatformCredential, UserPreference, MonthlySales, 
    SettlementDepartment, ServiceGroup, PurchaseGroup, Member, 
    PurchasePrice, MemberStat, TotalStat, PurchaseGroupAdUnit, ExchangeRate,
    MonthlyAdjustment
)
from ..platforms import get_platform_display_name, PLATFORM_ORDER

logger = logging.getLogger(__name__)

@login_required
def sales_report_view(request):
    """매출 현황 뷰 - 기본 조회 및 표시"""
    year = int(request.GET.get('year', date.today().year))

    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            return handle_excel_upload(request, year)
        elif request.POST.get('inline_edit') or 'save_changes' in request.POST:
            return handle_inline_edit(request, year)

    context = generate_sales_context(request, year)
    return render(request, 'sales.html', context)

@login_required
@api_view(['POST'])
def sales_api_view(request):
    """매출 관련 API 요청 처리"""
    data = request.data
    action = data.get('action')
    year = int(data.get('year', date.today().year))
    
    if action == 'delete':
        return handle_sales_delete(request, data, year)
    elif action == 'group':
        return handle_sales_group(request, data, year)
    elif action == 'get_groups':
        return handle_get_groups(request)
    elif action == 'ungroup':
        return handle_sales_ungroup(request, data, year)
    elif action == 'ungroup_multiple':
        return handle_sales_ungroup_multiple(request, data, year)

    return JsonResponse({'success': False, 'message': '알 수 없는 작업입니다.'}, status=400)

def handle_excel_upload(request, year):
    """엑셀 파일 업로드 처리"""
    excel_file = request.FILES['excel_file']
    try:
        print("--- 엑셀 파일 처리 시작 ---")
        if excel_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            nrows = sheet.max_row
        elif excel_file.name.endswith('.xls'):
            wb = xlrd.open_workbook(file_contents=excel_file.read())
            sheet = wb.sheet_by_index(0)
            nrows = sheet.nrows
        else:
            messages.error(request, "지원하지 않는 파일 형식입니다. .xlsx 또는 .xls 파일을 업로드해주세요.")
            return redirect(f'/sales-report/?year={year}')
        
        header_row_index = -1
        headers = []
        temp_rows_for_header = [sheet.row_values(i) for i in range(min(10, nrows))] if excel_file.name.endswith('.xls') else list(sheet.iter_rows(values_only=True, max_row=10))
        for i, row in enumerate(temp_rows_for_header):
            row_values = [str(cell) if cell is not None else '' for cell in row]
            if '작성일자' in row_values and '상호' in row_values and '공급가액' in row_values:
                header_row_index = i
                headers = [str(cell).strip() for cell in row]
                print(f"실제 헤더 발견 (행 {header_row_index + 1}): {headers}")
                break
        
        if header_row_index == -1:
            messages.error(request, "엑셀 파일에서 유효한 헤더를 찾을 수 없습니다. ('작성일자', '상호', '공급가액' 포함 필요)")
            return redirect(f'/sales-report/?year={year}')
        
        column_map = {header: i for i, header in enumerate(headers)}
        print(f"생성된 컬럼 맵: {column_map}")
        required_columns = ['상호', '품목명', '공급가액', '작성일자']
        if not all(col in column_map for col in required_columns):
            missing_cols = [col for col in required_columns if col not in column_map]
            messages.error(request, f"엑셀 파일에 필요한 컬럼이 없습니다: {', '.join(missing_cols)}")
            return redirect(f'/sales-report/?year={year}')
        
        # 사업자번호 컬럼 확인 (선택적)
        business_number_column = None
        if '공급받는자사업자등록번호' in column_map:
            business_number_column = column_map['공급받는자사업자등록번호']
            print(f"사업자번호 컬럼 발견: {business_number_column}번째 열")
        else:
            print("사업자번호 컬럼이 없습니다. 기본값으로 처리합니다.")
        
        created_count = 0
        updated_count = 0
        grouped_count = 0
        data_rows = (sheet.row_values(i) for i in range(header_row_index + 1, nrows)) if excel_file.name.endswith('.xls') else sheet.iter_rows(values_only=True, min_row=header_row_index + 2)
        
        # 사업자번호별 그룹화를 위한 임시 저장소
        business_number_groups = {}
        
        for i, row_data in enumerate(data_rows, start=header_row_index + 2):
            try:
                if not any(row_data) or not str(row_data[column_map['상호']]).strip():
                    continue
                company_name = str(row_data[column_map['상호']]).strip()
                service_name = str(row_data[column_map['품목명']]).strip()
                approval_number = str(row_data[column_map['승인번호']]).strip() if '승인번호' in column_map else f"{company_name}_{service_name}"
                
                # 사업자번호 추출
                business_number = ""
                if business_number_column is not None and len(row_data) > business_number_column:
                    business_number_raw = row_data[business_number_column]
                    if business_number_raw:
                        business_number = str(business_number_raw).strip()
                        # 사업자번호 형식 정리 (하이픈 제거 후 다시 추가)
                        business_number = business_number.replace('-', '').replace(' ', '')
                        if len(business_number) == 10:  # 10자리 숫자인 경우
                            business_number = f"{business_number[:3]}-{business_number[3:5]}-{business_number[5:]}"
                
                supply_value_raw = row_data[column_map['공급가액']]
                supply_value = float(str(supply_value_raw).replace(',', '')) if supply_value_raw else 0.0
                
                # 사업자번호별 그룹화 정보 수집
                if business_number:
                    if business_number not in business_number_groups:
                        business_number_groups[business_number] = {
                            'company_name': company_name,
                            'service_names': set(),
                            'service_codes': set(),
                            'total_amount': 0
                        }
                    business_number_groups[business_number]['service_names'].add(service_name)
                    business_number_groups[business_number]['service_codes'].add(approval_number)
                    business_number_groups[business_number]['total_amount'] += supply_value
                issue_date_raw = row_data[column_map['작성일자']]
                
                if isinstance(issue_date_raw, datetime):
                    issue_date = issue_date_raw.date()
                elif isinstance(issue_date_raw, str) and issue_date_raw:
                     issue_date = datetime.strptime(issue_date_raw, '%Y-%m-%d').date()
                elif isinstance(issue_date_raw, (int, float)):
                     issue_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(issue_date_raw) - 2).date()
                else:
                    print(f"[{i}번째 행] 날짜 형식 오류, 건너뜁니다: {issue_date_raw}")
                    continue
                
                if issue_date.year != year:
                     continue
                
                # MonthlySales에 직접 저장
                obj, created = MonthlySales.objects.get_or_create(
                    user=request.user,
                    service_code=approval_number,
                    year_month=issue_date.replace(day=1),
                    defaults={
                        'company_name': company_name,
                        'service_name': service_name,
                        'transaction_date': issue_date,
                        'amount': supply_value,
                        'business_number': business_number,
                    }
                )
                if not created:
                    obj.amount += supply_value
                    # 사업자번호가 있고 기존에 없었다면 업데이트
                    if business_number and not obj.business_number:
                        obj.business_number = business_number
                    obj.save()
                    updated_count += 1
                else:
                    created_count += 1
            except Exception as e:
                print(f"[오류] {i}번째 행 처리 실패: {row_data}, 원인: {e}")
                continue
        
        print(f"--- 엑셀 파일 처리 종료 ---")
        
        # 사업자번호별 자동 그룹화 수행
        if business_number_groups:
            print(f"사업자번호별 그룹화 시작: {len(business_number_groups)}개 그룹")
            for business_number, group_info in business_number_groups.items():
                try:
                    # 이미 해당 사업자번호로 그룹이 있는지 확인
                    existing_group = ServiceGroup.objects.filter(
                        user=request.user,
                        group_code=business_number
                    ).first()
                    
                    if existing_group:
                        # 기존 그룹이 있으면 해당 그룹에 연결
                        updated = MonthlySales.objects.filter(
                            user=request.user,
                            service_code__in=group_info['service_codes'],
                            year_month__year=year
                        ).update(group=existing_group)
                        grouped_count += updated
                        print(f"기존 그룹 '{existing_group.group_name}'에 {updated}개 항목 연결")
                    else:
                        # 새 그룹 생성
                        group_name = f"{group_info['company_name']} ({business_number})"
                        
                        # service_name 길이 제한 처리
                        service_names_list = sorted(group_info['service_names'])
                        if len(service_names_list) > 0:
                            # 첫 번째 서비스명을 사용하고, 여러 개인 경우 "외 N개" 표시
                            if len(service_names_list) == 1:
                                service_name = service_names_list[0]
                            else:
                                service_name = f"{service_names_list[0]} 외 {len(service_names_list)-1}개"
                        else:
                            service_name = "미정"
                        
                        new_group = ServiceGroup.objects.create(
                            user=request.user,
                            group_code=business_number,
                            group_name=group_name,
                            company_name=group_info['company_name'],
                            service_name=service_name
                        )
                        
                        # 해당 서비스 코드들을 새 그룹에 연결
                        updated = MonthlySales.objects.filter(
                            user=request.user,
                            service_code__in=group_info['service_codes'],
                            year_month__year=year
                        ).update(group=new_group)
                        grouped_count += updated
                        print(f"새 그룹 '{group_name}' 생성 및 {updated}개 항목 연결")
                        
                except Exception as e:
                    print(f"사업자번호 {business_number} 그룹화 중 오류: {e}")
                    continue
        
        messages.success(request, f"엑셀 파일 처리가 완료되었습니다. (신규: {created_count}건, 업데이트: {updated_count}건, 그룹화: {grouped_count}건)")
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f"엑셀 파일 처리 중 오류가 발생했습니다: {e}")
    
    return redirect(f"/sales-report/?year={year}")

def handle_inline_edit(request, year):
    """인라인 수정 처리"""
    updated_count = 0
    print("=== POST 데이터 디버깅 ===")
    for key, value in request.POST.items():
        print(f"POST key: {key}, value: {value}")
    print("=== POST 데이터 디버깅 끝 ===")
    
    for key, value in request.POST.items():
        try:
            if key.startswith('issue_type_') or key.startswith('settlement_department_') or key.startswith('settlement_period_') or key.startswith('code_') or key.startswith('company_name_') or key.startswith('service_name_'):
                # 필드명과 코드 분리 로직 개선
                if key.startswith('issue_type_') or key.startswith('settlement_department_') or key.startswith('settlement_period_'):
                    field_name = key.split('_')[0] + '_' + key.split('_')[1]  # 예: 'settlement_period'
                    code = key[len(field_name)+1:]  # 나머지 전체가 code (예: 'GROUP_20250622_141039')
                else:
                    field_name = '_'.join(key.split('_')[:-1])
                    code = key.split('_')[-1]
                print(f"처리 중: field_name={field_name}, code={code}, value={value}")
                
                # 그룹 수정 - 실제 그룹 존재 여부로 판단
                group = ServiceGroup.objects.filter(user=request.user, group_code=code).first()
                if group:
                    # 그룹코드 변경인지 확인
                    if field_name == 'code':
                        # 그룹코드 변경 시: POST key에서 현재 그룹코드 추출 (code_000001 -> 000001)
                        current_group_code = key.split('_', 1)[1]  # code_000001 -> 000001
                        group = ServiceGroup.objects.filter(user=request.user, group_code=current_group_code).first()
                        if not group:
                            print(f"그룹을 찾을 수 없음: {current_group_code}")
                            continue
                    else:
                        # 다른 필드 변경 시: 이미 위에서 찾은 그룹 사용
                        pass
                    
                    print(f"그룹 찾음: {group.group_name}")
                    
                    if field_name == 'issue_type':
                        print(f"이슈 타입 업데이트: {group.issue_type} -> {value}")
                        group.issue_type = value
                        group.save()
                        print(f"저장 완료: {group.issue_type}")
                    elif field_name == 'settlement_department':
                        print(f"정산부서 업데이트: {group.settlement_department} -> {value}")
                        try:
                            department = SettlementDepartment.objects.get(id=value, user=request.user) if value else None
                            group.settlement_department = department
                            group.save()
                            print(f"저장 완료: {group.settlement_department}")
                        except SettlementDepartment.DoesNotExist:
                            print(f"정산부서를 찾을 수 없음: {value}")
                            continue
                    elif field_name == 'settlement_period':
                        print(f"정산기간 업데이트: {group.settlement_timing} -> {value}")
                        group.settlement_timing = value
                        group.save()
                        print(f"저장 완료: {group.settlement_timing}")
                    elif field_name == 'code':
                        print(f"그룹코드 업데이트: {group.group_code} -> {value}")
                        # 그룹코드 중복 체크
                        if ServiceGroup.objects.filter(user=request.user, group_code=value).exclude(id=group.id).exists():
                            print(f"그룹코드 중복: {value}")
                            continue
                        group.group_code = value
                        group.save()
                        print(f"저장 완료: {group.group_code}")
                    elif field_name == 'company_name':
                        print(f"회사명 업데이트: {group.company_name} -> {value}")
                        group.company_name = value
                        group.save()
                        print(f"저장 완료: {group.company_name}")
                    elif field_name == 'service_name':
                        print(f"서비스명 업데이트: {group.service_name} -> {value}")
                        group.service_name = value
                        group.save()
                        print(f"저장 완료: {group.service_name}")
                    
                    updated_count += 1
                    print(f"업데이트 완료: {updated_count}")
                
                # 개별 항목 수정
                else:
                    try:
                        sales_item = MonthlySales.objects.get(id=code, user=request.user)
                        print(f"개별 항목 찾음: {sales_item}")
                        
                        if field_name == 'issue_type':
                            print(f"이슈 타입 업데이트: {sales_item.issue_type} -> {value}")
                            sales_item.issue_type = value
                            sales_item.save()
                            print(f"저장 완료: {sales_item.issue_type}")
                        elif field_name == 'settlement_department':
                            print(f"정산부서 업데이트: {sales_item.settlement_department} -> {value}")
                            try:
                                department = SettlementDepartment.objects.get(id=value, user=request.user) if value else None
                                sales_item.settlement_department = department
                                sales_item.save()
                                print(f"저장 완료: {sales_item.settlement_department}")
                            except SettlementDepartment.DoesNotExist:
                                print(f"정산부서를 찾을 수 없음: {value}")
                                continue
                        elif field_name == 'settlement_period':
                            print(f"정산기간 업데이트: {sales_item.settlement_timing} -> {value}")
                            sales_item.settlement_timing = value
                            sales_item.save()
                            print(f"저장 완료: {sales_item.settlement_timing}")
                        elif field_name == 'code':
                            print(f"서비스코드 업데이트: {sales_item.service_code} -> {value}")
                            sales_item.service_code = value
                            sales_item.save()
                            print(f"저장 완료: {sales_item.service_code}")
                        elif field_name == 'company_name':
                            print(f"회사명 업데이트: {sales_item.company_name} -> {value}")
                            sales_item.company_name = value
                            sales_item.save()
                            print(f"저장 완료: {sales_item.company_name}")
                        elif field_name == 'service_name':
                            print(f"서비스명 업데이트: {sales_item.service_name} -> {value}")
                            sales_item.service_name = value
                            sales_item.save()
                            print(f"저장 완료: {sales_item.service_name}")
                        
                        updated_count += 1
                        print(f"업데이트 완료: {updated_count}")
                    except MonthlySales.DoesNotExist:
                        print(f"개별 항목을 찾을 수 없음: {code}")
                        continue
        except (IndexError, ValueError) as e:
            print(f"키 파싱 오류: {key}, {e}")
            continue
        except Exception as e:
            print(f"업데이트 중 오류 발생: {e}")
            continue
    
    if updated_count > 0:
        messages.success(request, f"{updated_count}개 항목이 성공적으로 업데이트되었습니다.")
    
    return redirect(f'/sales-report/?year={year}')

def handle_sales_delete(request, data, year):
    """매출 데이터 삭제 처리"""
    codes = data.get('codes', [])
    if not codes:
        return JsonResponse({'success': False, 'message': '삭제할 항목이 선택되지 않았습니다.'})
    
    try:
        deleted_count = 0
        deleted_groups = 0
        
        for code in codes:
            # 그룹 코드로 실제 그룹이 존재하는지 먼저 확인
            group = ServiceGroup.objects.filter(group_code=code, user=request.user).first()
            if group:
                # 그룹에 속한 MonthlySales 삭제
                count = MonthlySales.objects.filter(
                    group=group,
                    user=request.user,
                    year_month__year=year
                ).delete()[0]
                deleted_count += count
                
                # 그룹 자체 삭제
                group.delete()
                deleted_groups += 1
            else:
                # 개별 항목 삭제
                count = MonthlySales.objects.filter(
                    service_code=code,
                    user=request.user,
                    year_month__year=year
                ).delete()[0]
                deleted_count += count
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count}개 항목과 {deleted_groups}개 그룹이 삭제되었습니다.',
            'deleted_count': deleted_count,
            'deleted_groups': deleted_groups
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'삭제 중 오류가 발생했습니다: {str(e)}'
        })

def handle_sales_group(request, data, year):
    """매출 데이터 그룹화 처리"""
    codes = data.get('codes', [])
    group_name = data.get('groupName', '').strip()  # JavaScript에서 groupName으로 보냄
    existing_group_code = data.get('existingGroupCode')
    
    if not codes:
        return JsonResponse({'success': False, 'message': '그룹화할 항목이 선택되지 않았습니다.'})
    
    if not group_name and not existing_group_code:
        return JsonResponse({'success': False, 'message': '그룹명을 입력해주세요.'})
    
    try:
        # 기존 그룹에 추가하는 경우
        if existing_group_code:
            group = ServiceGroup.objects.get(
                group_code=existing_group_code,
                user=request.user
            )
            group_name = group.group_name
        else:
            # 새 그룹 생성
            if not group_name:
                return JsonResponse({'success': False, 'message': '그룹명을 입력해주세요.'})
            
            # 선택된 코드들 중 그룹 코드가 있는지 확인하고, 있다면 먼저 해제 및 삭제
            group_codes_to_ungroup = []
            individual_codes = []
            
            for code in codes:
                # 그룹 코드인지 확인 - 실제 그룹 존재 여부로 판단
                if ServiceGroup.objects.filter(user=request.user, group_code=code).exists():
                    group_codes_to_ungroup.append(code)
                else:
                    individual_codes.append(code)
            
            # 그룹 해제 및 삭제 작업 수행
            ungrouped_count = 0
            deleted_groups = 0
            for group_code in group_codes_to_ungroup:
                try:
                    existing_group = ServiceGroup.objects.get(
                        group_code=group_code,
                        user=request.user
                    )
                    count = MonthlySales.objects.filter(
                        group=existing_group,
                        user=request.user,
                        year_month__year=year
                    ).update(group=None)
                    ungrouped_count += count
                    
                    # 그룹 자체 삭제
                    existing_group.delete()
                    deleted_groups += 1
                    
                except ServiceGroup.DoesNotExist:
                    continue
            
            # 새로운 그룹 생성
            group_code = generate_group_code()
            group = ServiceGroup.objects.create(
                user=request.user,
                group_code=group_code,
                group_name=group_name,
                company_name=group_name,
                service_name=group_name
            )
            
            # 모든 코드(해제된 그룹의 개별 항목들 + 원래 개별 항목들)를 새 그룹에 연결
            all_codes = individual_codes.copy()
            # 해제된 그룹에 속했던 항목들의 service_code를 찾아서 추가
            for group_code in group_codes_to_ungroup:
                # 해당 그룹에 속했던 MonthlySales의 service_code들을 가져옴 (이미 group=None으로 설정됨)
                service_codes = MonthlySales.objects.filter(
                    user=request.user,
                    year_month__year=year,
                    group__isnull=True  # 그룹이 해제된 항목들
                ).values_list('service_code', flat=True).distinct()
                all_codes.extend(service_codes)
            
            # 중복 제거
            all_codes = list(set(all_codes))
            
            # 선택된 항목들을 그룹에 연결
            updated_count = MonthlySales.objects.filter(
                service_code__in=all_codes,
                user=request.user,
                year_month__year=year
            ).update(group=group)
            
            return JsonResponse({
                'success': True,
                'message': f'{deleted_groups}개 그룹이 삭제되고, {updated_count}개 항목이 새 그룹 "{group_name}"에 추가되었습니다.',
                'group_code': group.group_code
            })
        
        # 기존 그룹에 추가하는 경우 (기존 로직)
        # 선택된 항목들을 그룹에 연결 (service_code로 찾기)
        updated_count = MonthlySales.objects.filter(
            service_code__in=codes,
            user=request.user,
            year_month__year=year
        ).update(group=group)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count}개 항목이 그룹 "{group_name}"에 추가되었습니다.',
            'group_code': group.group_code
        })
        
    except ServiceGroup.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '선택한 그룹을 찾을 수 없습니다.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'그룹화 중 오류가 발생했습니다: {str(e)}'
        })

def handle_get_groups(request):
    """사용 가능한 그룹 목록 조회"""
    try:
        groups = ServiceGroup.objects.filter(user=request.user).values('group_code', 'group_name')
        return JsonResponse({
            'success': True,
            'groups': list(groups)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'그룹 목록 조회 중 오류가 발생했습니다: {str(e)}'
        })

def handle_sales_ungroup(request, data, year):
    """매출 데이터 그룹 해제 처리"""
    codes = data.get('codes', [])
    
    if not codes:
        return JsonResponse({'success': False, 'message': '그룹 해제할 항목이 선택되지 않았습니다.'})
    
    try:
        updated_count = MonthlySales.objects.filter(
            service_code__in=codes,
            user=request.user,
            year_month__year=year
        ).update(group=None)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count}개 항목의 그룹이 해제되었습니다.',
            'ungrouped_count': updated_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'그룹 해제 중 오류가 발생했습니다: {str(e)}'
        })

def handle_sales_ungroup_multiple(request, data, year):
    """매출 데이터 다중 그룹 해제 처리"""
    codes = data.get('codes', [])
    
    if not codes:
        return JsonResponse({'success': False, 'message': '그룹 해제할 항목이 선택되지 않았습니다.'})
    
    try:
        # 그룹 코드로 그룹을 찾고, 해당 그룹에 속한 모든 MonthlySales의 그룹을 해제한 후 그룹 삭제
        updated_count = 0
        deleted_groups = 0
        
        for group_code in codes:
            try:
                group = ServiceGroup.objects.get(
                    group_code=group_code,
                    user=request.user
                )
                # 해당 그룹에 속한 MonthlySales의 그룹을 해제
                count = MonthlySales.objects.filter(
                    group=group,
                    user=request.user,
                    year_month__year=year
                ).update(group=None)
                updated_count += count
                
                # 그룹 자체 삭제
                group.delete()
                deleted_groups += 1
                
            except ServiceGroup.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_groups}개 그룹이 삭제되고 {updated_count}개 항목의 그룹이 해제되었습니다.',
            'ungrouped_count': updated_count,
            'deleted_groups': deleted_groups
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'그룹 해제 중 오류가 발생했습니다: {str(e)}'
        })

def generate_sales_context(request, year):
    """매출 현황 컨텍스트 생성"""
    # 년월 검색 파라미터 처리 (home.html과 동일한 방식)
    adjustment_month = request.GET.get('adjustment_month')
    if adjustment_month:
        try:
            # YYYY-MM 형식에서 년도와 월 추출
            selected_date = datetime.strptime(adjustment_month, '%Y-%m').date()
            adjustment_year = selected_date.year
            adjustment_month = selected_date.month
        except ValueError:
            adjustment_year = year
            adjustment_month = date.today().month
    else:
        adjustment_year = year
        adjustment_month = date.today().month
    
    # 발행유형 선택지 추가
    issue_type_choices = [
        ('정발행', '정발행'),
        ('역발행', '역발행'),
        ('영세율', '영세율'),
    ]
    
    # 월별 데이터 조회 (그룹화 전 원본 데이터)
    monthly_data = MonthlySales.objects.filter(
        user=request.user,
        year_month__year=year
    ).order_by('year_month', 'company_name', 'service_name')
    
    # 1. 월별 매출/매입 계산 (캐시 적용)
    cache_key = f"monthly_totals_{request.user.id}_{year}"
    cached_monthly_data = cache.get(cache_key)
    
    if cached_monthly_data is None:
        # 캐시가 없으면 계산
        logger.info(f"캐시 MISS: {cache_key} - 데이터 계산 시작")
        monthly_totals = {m: 0 for m in range(1, 13)}  # 매출 (양수)
        purchase_monthly = {m: 0 for m in range(1, 13)}  # 매입 (음수 절대값)
        
        for month in range(1, 13):
            month_start = date(year, month, 1)
            
            # 매출 데이터 (양수)
            revenue = monthly_data.filter(
                year_month=month_start,
                amount__gt=0
            ).aggregate(
                total_amount=Sum('amount')
            )['total_amount'] or 0
            
            # 매입 데이터 (음수 절대값)
            purchase = monthly_data.filter(
                year_month=month_start,
                amount__lt=0
            ).aggregate(
                total_amount=Sum('amount')
            )['total_amount'] or 0
            
            monthly_totals[month] = revenue
            purchase_monthly[month] = abs(purchase)
        
        # 24시간 캐시 저장 (86400초)
        cache.set(cache_key, {
            'monthly_totals': monthly_totals,
            'purchase_monthly': purchase_monthly
        }, 86400)
        logger.info(f"캐시 SET: {cache_key} - 데이터 저장 완료")
    else:
        # 캐시에서 데이터 가져오기
        logger.info(f"캐시 HIT: {cache_key} - 캐시된 데이터 사용")
        monthly_totals = cached_monthly_data['monthly_totals']
        purchase_monthly = cached_monthly_data['purchase_monthly']
    
    # 전월대비 증감 계산
    monthly_changes = _calculate_monthly_changes(monthly_totals)
    
    # 2. 그룹별 데이터 (캐시 적용)
    groups = ServiceGroup.objects.filter(user=request.user).order_by('group_name')
    departments = SettlementDepartment.objects.filter(user=request.user).order_by('name')
    
    # sales_data 생성 - 그룹별 및 개별 데이터 (표시용)
    sales_data = []
    
    # 그룹별 데이터 캐시 확인
    group_cache_key = f"group_sales_data_{request.user.id}_{year}"
    cached_group_data = cache.get(group_cache_key)
    
    if cached_group_data is None:
        # 캐시가 없으면 계산
        logger.info(f"그룹 데이터 캐시 MISS: {group_cache_key} - 데이터 계산 시작")
        for group in groups:
            group_sales = monthly_data.filter(group=group)
            monthly_sales = {}
            monthly_changes = {}
            total_amount = 0
            
            for month in range(1, 13):
                month_start = date(year, month, 1)
                month_amount = group_sales.filter(year_month=month_start).aggregate(
                    total=Sum('amount')
                )['total'] or 0
                monthly_sales[month] = month_amount
                total_amount += month_amount
                
                # 증감 계산
                if month == 1:
                    monthly_changes[month] = month_amount
                else:
                    prev_month_amount = monthly_sales.get(month - 1, 0)
                    monthly_changes[month] = month_amount - prev_month_amount
            
            sales_data.append({
                'is_group': True,
                'code': group.group_code,
                'company_name': group.company_name,
                'service_name': group.service_name,
                'issue_type': group.issue_type or '',
                'settlement_department': group.settlement_department.name if group.settlement_department else '',
                'settlement_period': group.settlement_timing or '',
                'monthly_sales': monthly_sales,
                'monthly_changes': monthly_changes,
                'total_amount': total_amount,
                'raw_issue_type': group.issue_type,
                'raw_department_id': group.settlement_department.id if group.settlement_department else None,
                'settlement_timing': group.settlement_timing or '',
                'service_mapping': {},
                'service_names_list': [],
            })
        
        # 24시간 캐시 저장 (86400초)
        cache.set(group_cache_key, sales_data, 86400)
        logger.info(f"그룹 데이터 캐시 SET: {group_cache_key} - 데이터 저장 완료")
    else:
        # 캐시에서 데이터 가져오기
        logger.info(f"그룹 데이터 캐시 HIT: {group_cache_key} - 캐시된 데이터 사용")
        sales_data = cached_group_data
    
    # 그룹에 속하지 않은 개별 데이터 처리
    ungrouped_sales = monthly_data.filter(group__isnull=True)
    service_groups = {}
    for sale in ungrouped_sales:
        service_code = sale.service_code
        if service_code not in service_groups:
            service_groups[service_code] = []
        service_groups[service_code].append(sale)
    
    for service_code, sales_list in service_groups.items():
        monthly_sales = {}
        monthly_changes = {}
        total_amount = 0
        
        for month in range(1, 13):
            month_start = date(year, month, 1)
            month_amount = sum(sale.amount for sale in sales_list if sale.year_month == month_start)
            monthly_sales[month] = month_amount
            total_amount += month_amount
            
            # 증감 계산
            if month == 1:
                monthly_changes[month] = month_amount
            else:
                prev_month_amount = monthly_sales.get(month - 1, 0)
                monthly_changes[month] = month_amount - prev_month_amount
        
        first_sale = sales_list[0]
        sales_data.append({
            'is_group': False,
            'code': service_code,
            'company_name': first_sale.company_name,
            'service_name': first_sale.service_name,
            'issue_type': '',
            'settlement_department': '',
            'settlement_period': '',
            'monthly_sales': monthly_sales,
            'monthly_changes': monthly_changes,
            'total_amount': total_amount,
            'raw_issue_type': '',
            'raw_department_id': None,
            'settlement_timing': '',
            'service_mapping': {},
            'service_names_list': [first_sale.service_name],
        })
    
    # 3. 유형별 통계 (표시용)
    type_summaries = []
    type_data = {}
    
    # 기본 발행유형들을 미리 초기화
    for value, display in issue_type_choices:
        type_data[value] = {
            'name': display,
            'count': 0,
            'monthly_amounts': {m: 0 for m in range(1, 13)},
            'total_amount': 0
        }
    
    # 실제 데이터로 통계 계산
    for item in sales_data:
        issue_type = item['issue_type'] or '미분류'
        if issue_type not in type_data:
            type_data[issue_type] = {
                'name': issue_type,
                'count': 0,
                'monthly_amounts': {m: 0 for m in range(1, 13)},
                'total_amount': 0
            }
        
        type_data[issue_type]['count'] += 1
        type_data[issue_type]['total_amount'] += item['total_amount']
        
        for month in range(1, 13):
            type_data[issue_type]['monthly_amounts'][month] += item['monthly_sales'].get(month, 0)
    
    type_summaries = list(type_data.values())
    
    # 4. 최종 계산
    # revenue_monthly에 매출(양수) + 매입(음수) 합계 포함
    revenue_monthly = {}
    for month in range(1, 13):
        revenue = monthly_totals.get(month, 0)  # 매출 (양수)
        purchase = purchase_monthly.get(month, 0)  # 매입 (음수 절대값)
        # 매출 + 매입(음수) = 매출 - 매입 절대값
        revenue_monthly[month] = revenue - purchase
    
    # 주요퍼블리셔만 보기 조건일 때의 매입 데이터 계산 (캐시 적용)
    purchase_cache_key = f"purchase_monthly_{request.user.id}_{year}"
    cached_purchase_data = cache.get(purchase_cache_key)
    
    if cached_purchase_data is None:
        # 캐시가 없으면 계산
        logger.info(f"매입 데이터 캐시 MISS: {purchase_cache_key} - 데이터 계산 시작")
        purchase_monthly = {m: 0 for m in range(1, 13)}
        
        # 주요 퍼블리셔 그룹 조회
        important_groups = PurchaseGroup.objects.filter(
            user=request.user, 
            is_active=True,
            is_important=True
        ).prefetch_related('ad_units')
        
        if important_groups.exists():
            # 모든 광고 단위 ID 수집
            all_ad_unit_ids = []
            for group in important_groups:
                for ad_unit in group.ad_units.all():
                    if ad_unit.is_active:
                        all_ad_unit_ids.append(ad_unit.ad_unit_id)
            
            # AdStats 데이터 일괄 조회
            ad_stats = AdStats.objects.filter(
                user=request.user,
                ad_unit_id__in=all_ad_unit_ids,
                date__year=year
            ).values('date', 'platform', 'ad_unit_id').annotate(
                earnings=Sum('earnings'),
                earnings_usd=Sum('earnings_usd')
            )
            
            # 환율 데이터 일괄 조회
            exchange_rates = ExchangeRate.objects.filter(
                user=request.user, 
                year_month__gte=date(year, 1, 1)
            ).values('year_month', 'usd_to_krw')
            exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}
            
            # stats_map 생성
            stats_map = {}
            for stat in ad_stats:
                key = (stat['date'], stat['ad_unit_id'])
                if stat['platform'] == 'adsense':
                    month_start = stat['date'].replace(day=1)
                    exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
                    stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
                elif stat['platform'] in ['admanager', 'adpost']:
                    stats_map[key] = stat['earnings'] or 0
                else:
                    stats_map[key] = stat['earnings'] or 0
        
        # 파워링크 데이터 조회
        adpost_stats = AdStats.objects.filter(
            platform='adpost',
            ad_unit_id='모바일뉴스픽_컨텐츠',
            date__year=year,
            user=request.user
        ).values('date').annotate(
            earnings=Sum('earnings'),
            clicks=Sum('clicks')
        )
        adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
        
        total_powerlink_stats = TotalStat.newspic_objects().filter(
            sdate__year=year
        ).values('sdate').annotate(
            total_powerlink=Sum('powerlink_count')
        )
        total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
        
        publisher_keys = [group.member.request_key for group in important_groups]
        member_powerlink_stats = TotalStat.newspic_objects().filter(
            request_key__in=publisher_keys,
            sdate__year=year
        ).values('request_key', 'sdate').annotate(
            powerlink_count=Sum('powerlink_count'),
            click_count=Sum('click_count')
        )
        member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}
        
        # 각 그룹별로 월별 매입비용 계산
        for group in important_groups:
            monthly_cost = {m: 0 for m in range(1, 13)}
            default_price = group.default_unit_price
            default_type = group.default_unit_type
            ad_units = group.ad_units.filter(is_active=True)
            
            for m in range(1, 13):
                last_day = calendar.monthrange(year, m)[1]
                for d in range(1, last_day + 1):
                    current_date = date(year, m, d)
                    
                    # AdSense/AdManager 일별 수익 계산
                    ad_revenue = Decimal('0')
                    for ad_unit in ad_units:
                        stat_key = (current_date, ad_unit.ad_unit_id)
                        ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
                    
                    # AdPost 일별 수익 분배
                    adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                    total_powerlink_count = total_powerlink_data.get(current_date, 0)
                    member_data = member_powerlink_data.get((group.member.request_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                    powerlink_count = member_data['powerlink_count']
                    click_count = member_data['click_count']
                    
                    adpost_earnings = adpost_info['earnings']
                    adpost_clicks = adpost_info['clicks']
                    powerlink_revenue = Decimal('0')
                    
                    if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                        unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                    
                    ad_revenue += powerlink_revenue
                    
                    # 매입비용 계산
                    rs_rate = default_price or 0
                    rs_type = default_type or 'percent'
                    purchase_cost = Decimal('0')
                    
                    if rs_type == 'percent':
                        purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
                    else:
                        purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
                    
                    # 월별 매입비용에 누적
                    monthly_cost[m] += int(purchase_cost)
            
            # 퍼블리셔 레벨에 따라 purchase_monthly에 누적
            if group.member.level == 50:  # 퍼블리셔
                for m in range(1, 13):
                    purchase_monthly[m] += monthly_cost[m]
        
        # 24시간 캐시 저장 (86400초)
        cache.set(purchase_cache_key, purchase_monthly, 86400)
        logger.info(f"매입 데이터 캐시 SET: {purchase_cache_key} - 데이터 저장 완료")
    else:
        # 캐시에서 데이터 가져오기
        logger.info(f"매입 데이터 캐시 HIT: {purchase_cache_key} - 캐시된 데이터 사용")
        purchase_monthly = cached_purchase_data
    
    gross_profit_monthly = {m: revenue_monthly.get(m, 0) - purchase_monthly.get(m, 0) for m in range(1, 13)}
    
    # 증감 데이터 생성
    revenue_changes = _calculate_monthly_changes(revenue_monthly)
    purchase_changes = {m: 0 for m in range(1, 13)}  # 매입 증감은 별도 계산 필요
    gross_profit_changes = {m: revenue_changes.get(m, 0) - purchase_changes.get(m, 0) for m in range(1, 13)}
    
    # 총합 계산
    total_revenue = sum(revenue_monthly.values())
    total_purchase = sum(purchase_monthly.values())
    total_gross_profit = sum(gross_profit_monthly.values())
    
    # 이익율 계산
    profit_rate_monthly = {}
    for month in range(1, 13):
        revenue = revenue_monthly.get(month, 0)
        if revenue > 0:
            profit_rate_monthly[month] = (gross_profit_monthly.get(month, 0) / revenue) * 100
        else:
            profit_rate_monthly[month] = 0
    
    total_profit_rate = (total_gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # 현재 년월 정보
    current_year = date.today().year
    current_month = date.today().month
    is_current_year = year == current_year
    
    # 매출정산 표 데이터 추가 - 검색된 년월 사용
    selected_year = int(adjustment_year)
    selected_month = int(adjustment_month)
    
    current_month_start = date(selected_year, selected_month, 1)
    prev_month_start = date(selected_year, selected_month-1, 1) if selected_month > 1 else date(selected_year-1, 12, 1)
    
    # 현재월 매출 조정 데이터 조회
    current_sales_adjustments = MonthlyAdjustment.objects.filter(
        user=request.user,
        year_month=current_month_start,
        adjustment_type='sales'
    ).values('sales_type', 'adjustment_amount', 'adjustment_note')
    
    # 전월 매출 조정 데이터 조회
    prev_sales_adjustments = MonthlyAdjustment.objects.filter(
        user=request.user,
        year_month=prev_month_start,
        adjustment_type='sales'
    ).values('sales_type', 'adjustment_amount', 'adjustment_note')
    
    # 매출정산 표 데이터 구성
    sales_types = ['영세율매출', '역발행매출', '정발행매출']
    sales_summary_rows = []
    sales_summary_prev_rows = []
    
    # 업체 수 계산을 위한 데이터 조회
    current_service_groups = ServiceGroup.objects.filter(
        user=request.user,
        is_active=True
    ).values('issue_type')
    
    current_company_counts = {}
    for group in current_service_groups:
        issue_type = group['issue_type']
        if issue_type == '영세율':
            current_company_counts['영세율매출'] = current_company_counts.get('영세율매출', 0) + 1
        elif issue_type == '역발행':
            current_company_counts['역발행매출'] = current_company_counts.get('역발행매출', 0) + 1
        elif issue_type == '정발행':
            current_company_counts['정발행매출'] = current_company_counts.get('정발행매출', 0) + 1
    
    # 각 발행유형별 매출 데이터 계산
    for sales_type in sales_types:
        # 발행유형 매핑
        issue_type_map = {
            '영세율매출': '영세율',
            '역발행매출': '역발행', 
            '정발행매출': '정발행'
        }
        issue_type = issue_type_map[sales_type]
        
        # 현재월 해당 발행유형 매출 계산
        current_month_sales = MonthlySales.objects.filter(
            user=request.user,
            group__issue_type=issue_type,
            group__is_active=True,
            year_month=current_month_start
        ).aggregate(
            total_sales=Sum('amount')
        )['total_sales'] or 0
        
        # 전월 해당 발행유형 매출 계산
        prev_month_sales = MonthlySales.objects.filter(
            user=request.user,
            group__issue_type=issue_type,
            group__is_active=True,
            year_month=prev_month_start
        ).aggregate(
            total_sales=Sum('amount')
        )['total_sales'] or 0
        
        # 현재월 데이터
        current_data = next((item for item in current_sales_adjustments if item['sales_type'] == sales_type), None)
        
        sales_summary_rows.append({
            'type': sales_type,
            'company_count': current_company_counts.get(sales_type, 0),
            'estimate': int(current_month_sales),  # 해당 발행유형의 실제 매출
            'note': current_data['adjustment_note'] if current_data else '(수기입력)'
        })
        
        # 전월 데이터
        prev_data = next((item for item in prev_sales_adjustments if item['sales_type'] == sales_type), None)
        prev_adjust_amount = int(prev_data['adjustment_amount']) if prev_data else 0
        
        sales_summary_prev_rows.append({
            'type': sales_type,
            'estimate': int(prev_month_sales),  # 해당 발행유형의 실제 매출
            'confirmed': int(prev_month_sales + prev_adjust_amount),  # 조정 전 + 조정금액
            'adjust': prev_adjust_amount,  # 조정금액
            'note': prev_data['adjustment_note'] if prev_data else '(수기입력)'
        })
    
    # 합계 계산
    sales_summary_total = {
        'company_count': sum(r.get('company_count', 0) for r in sales_summary_rows),
        'estimate': int(sum(r.get('estimate', 0) or 0 for r in sales_summary_rows))
    }
    
    sales_summary_prev_total = {
        'estimate': int(sum(r.get('estimate', 0) or 0 for r in sales_summary_prev_rows)),  # 조정 전 매출추정액 합계 (조정금액 제외)
        'confirmed': int(sum(r.get('confirmed', 0) or 0 for r in sales_summary_prev_rows)),  # 조정 후 매출확정액 합계
        'adjust': int(sum(r.get('adjust', 0) or 0 for r in sales_summary_prev_rows))  # 조정금액 합계
    }
    
    # 전월 정보
    prev_year = selected_year if selected_month > 1 else selected_year - 1
    prev_mon = selected_month - 1 if selected_month > 1 else 12
    prev_month = f"{prev_year:04d}-{prev_mon:02d}"
    selected_month = f"{selected_year:04d}-{selected_month:02d}"
    
    context = {
        'year': year,
        'monthly_data': monthly_data,
        'monthly_totals': monthly_totals,
        'monthly_changes': monthly_changes,
        'groups': groups,
        'departments': departments,
        'months': list(range(1, 13)),
        'sales_data': sales_data,
        'type_summaries': type_summaries,
        'revenue_monthly': revenue_monthly,
        'purchase_monthly': purchase_monthly,
        'gross_profit_monthly': gross_profit_monthly,
        'revenue_changes': revenue_changes,
        'purchase_changes': purchase_changes,
        'gross_profit_changes': gross_profit_changes,
        'total_revenue': total_revenue,
        'total_purchase': total_purchase,
        'total_gross_profit': total_gross_profit,
        'profit_rate_monthly': profit_rate_monthly,
        'total_profit_rate': total_profit_rate,
        'is_current_year': is_current_year,
        'current_month': current_month,
        'issue_type_choices': issue_type_choices,
        # 매출정산 표용 context
        'sales_summary_prev_rows': sales_summary_prev_rows,
        'sales_summary_rows': sales_summary_rows,
        'sales_summary_total': sales_summary_total,
        'sales_summary_prev_total': sales_summary_prev_total,
        'selected_month': selected_month,
        'prev_month': prev_month,
        # 년월 검색용 context (home.html과 동일한 방식)
        'selected_adjustment_month': f"{adjustment_year:04d}-{adjustment_month:02d}",
    }
    
    return context

def generate_group_code():
    """그룹 코드 생성"""
    import random
    import string
    
    while True:
        # 6자리 랜덤 코드 생성
        code = ''.join(random.choices(string.digits, k=6))
        
        # 중복 체크
        if not ServiceGroup.objects.filter(group_code=code).exists():
            return code

def _calculate_monthly_changes(monthly_amounts):
    """월별 증감 계산"""
    changes = {}
    for month in range(1, 13):
        current = monthly_amounts.get(month, 0)
        if month == 1:
            prev_month = 12
            prev_year = True
        else:
            prev_month = month - 1
            prev_year = False
        
        # 이전 월 데이터는 현재 모델에서 조회 불가하므로 0으로 설정
        previous = 0
        
        if previous > 0:
            change_amount = current - previous
        else:
            change_amount = current
        
        changes[month] = change_amount
    
    return changes 