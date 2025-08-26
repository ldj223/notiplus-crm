import os, json
from datetime import datetime, timedelta, date
from django.conf import settings
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count
from django.contrib import messages
from django.utils import timezone
import openpyxl
import xlrd
from django.core.exceptions import ValidationError
from decimal import Decimal, ROUND_HALF_UP
import logging
from rest_framework.decorators import api_view
from django.db.models import Q
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import calendar
from django.core.cache import cache

from ..forms import CredentialForm, SignUpForm
from ..models import (
    AdStats, PlatformCredential, UserPreference, MonthlySales, 
    SettlementDepartment, ServiceGroup, PurchaseGroup, Member, 
    PurchasePrice, MemberStat, TotalStat, PurchaseGroupAdUnit, ExchangeRate,
    MonthlyAdjustment
)
from ..platforms import get_platform_display_name, PLATFORM_ORDER

logger = logging.getLogger(__name__)

def calculate_purchase_cost_by_date_range(user, start_date, end_date):
    """
    지정된 날짜 범위의 매입비용을 계산하여 반환
    Returns: {
        'publisher_cost': {date: cost, ...},
        'partners_cost': {date: cost, ...}
    }
    """
    from decimal import Decimal, ROUND_HALF_UP
    
    # 날짜 리스트 생성
    date_list = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    
    # 결과 초기화
    result = {
        'publisher_cost': {d: Decimal('0') for d in date_list},
        'partners_cost': {d: Decimal('0') for d in date_list}
    }
    
    # 1. 모든 PurchaseGroup 조회
    all_groups = PurchaseGroup.objects.filter(user=user, is_active=True)
    
    # 2. 모든 광고 단위 ID 수집
    all_ad_unit_ids = []
    for group in all_groups:
        for ad_unit in group.ad_units.all():
            if ad_unit.is_active:
                all_ad_unit_ids.append(ad_unit.ad_unit_id)
    
    # 3. AdStats 데이터 일괄 조회 (세부 데이터와 동일한 방식)
    ad_stats = AdStats.objects.filter(
        user=user,
        date__range=[start_date, end_date],
        ad_unit_id__in=all_ad_unit_ids
    ).values('date', 'platform', 'ad_unit_id').annotate(
        earnings=Sum('earnings'),
        earnings_usd=Sum('earnings_usd')
    )
    
    # 4. 환율 데이터 조회 (세부 데이터와 동일한 방식)
    exchange_rates = ExchangeRate.objects.filter(
        user=user, 
        year_month__gte=start_date.replace(day=1)
    ).values('year_month', 'usd_to_krw')
    exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}
    
    # 5. stats_map 생성 (세부 데이터와 동일한 방식)
    stats_map = {}
    for stat in ad_stats:
        key = (stat['date'], stat['ad_unit_id'])
        if stat['platform'] == 'adsense':
            month_start = stat['date'].replace(day=1)
            exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
            stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
        elif stat['platform'] == 'admanager':
            # 애드매니저(ADX)는 KRW 그대로
            stats_map[key] = stat['earnings'] or 0
        elif stat['platform'] == 'adpost':
            # 애드포스트는 KRW 그대로
            stats_map[key] = stat['earnings'] or 0
        else:
            # 기타 플랫폼은 KRW 그대로
            stats_map[key] = stat['earnings'] or 0
    
    # 6. 파워링크(애드포스트) 수익 분배 로직 (세부 데이터와 동일한 방식)
    # 6-1. 파워링크(애드포스트) 데이터 일괄 조회
    adpost_stats = AdStats.objects.filter(
        platform='adpost',
        ad_unit_id='모바일뉴스픽_컨텐츠',
        date__range=[start_date, end_date],
        user=user
    ).values('date').annotate(
        earnings=Sum('earnings'),
        clicks=Sum('clicks')
    )
    adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
    
    # 6-2. 전체 파워링크 클릭수 일괄 조회
    total_powerlink_stats = TotalStat.newspic_objects().filter(
        sdate__range=[start_date, end_date]
    ).values('sdate').annotate(
        total_powerlink=Sum('powerlink_count')
    )
    total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
    
    # 6-3. 퍼블리셔별 파워링크 클릭수 일괄 조회
    publisher_keys = [group.member.request_key for group in all_groups]
    member_powerlink_stats = TotalStat.newspic_objects().filter(
        request_key__in=publisher_keys,
        sdate__range=[start_date, end_date]
    ).values('request_key', 'sdate').annotate(
        powerlink_count=Sum('powerlink_count'),
        click_count=Sum('click_count')
    )
    member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}
    
    # 7. 각 그룹별 일별 매입비용 계산
    for group in all_groups:
        member_level = group.member.level if group.member else 60
        default_price = group.default_unit_price
        default_type = group.default_unit_type
        
        # 해당 그룹의 매핑된 광고 단위들 조회
        ad_units = group.ad_units.filter(is_active=True)
        
        for current_date in date_list:
            ad_revenue = Decimal('0')
            
            # 1. 애드센스/애드매니저 광고수익 합산
            for ad_unit in ad_units:
                stat_key = (current_date, ad_unit.ad_unit_id)
                ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
            
            # 2. 파워링크(애드포스트) 수익 분배
            adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
            total_powerlink_count = total_powerlink_data.get(current_date, 0)
            member_data = member_powerlink_data.get((group.member.request_key, current_date), {'powerlink_count': 0, 'click_count': 0})
            powerlink_count = member_data['powerlink_count']
            click_count = member_data['click_count']
            
            adpost_earnings = adpost_info['earnings']
            adpost_clicks = adpost_info['clicks']
            powerlink_revenue = Decimal('0')
            
            if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
            
            ad_revenue += powerlink_revenue
            
            # 3. 매입비용 계산
            rs_rate = default_price or 0
            rs_type = default_type or 'percent'
            purchase_cost = Decimal('0')
            
            if rs_type == 'percent':
                purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
            else:
                # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
            
            # 4. level에 따라 매입비용 분류
            if member_level == 50:  # 퍼블리셔
                result['publisher_cost'][current_date] += int(purchase_cost)
            elif member_level in [60, 61, 65]:  # 파트너스
                # 파트너스는 기존 방식으로 계산하지 않음 (유효PV * 5로 계산)
                pass
    
    # 8. 파트너스 비용을 유효PV * 5로 계산
    # 순환 import 방지를 위해 함수 내부에서 import
    import importlib
    reports_module = importlib.import_module('stats.views.reports')
    partners_valid_pv_data = reports_module.get_partners_valid_pv_data(user, start_date, end_date)
    
    for current_date in date_list:
        if current_date in partners_valid_pv_data['daily']:
            valid_pv = partners_valid_pv_data['daily'][current_date]['valid_pageview']
            # 파트너스 비용 = 유효PV * 5
            result['partners_cost'][current_date] = int(valid_pv * 5)
    
    return result

@login_required
def purchase_report_view(request): 
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', date.today().month))
    search_query = request.GET.get('search', '').strip()
    important_only = request.GET.get('important_only', '1') == '1'
    
    # 요청 추적 로그
    logger.info(f"요청 시작: User={request.user.id}, Year={year}, Search={search_query}, Important={important_only}, Session={request.session.session_key}")
    
    # 간단한 캐시 테스트
    test_cache_key = f"test_cache_{request.user.id}"
    test_data = cache.get(test_cache_key)
    if test_data:
        logger.info(f"테스트 캐시 HIT: {test_data}")
    else:
        test_data = f"테스트 데이터 {datetime.now()}"
        cache.set(test_cache_key, test_data, 60)  # 1분 캐시
        logger.info(f"테스트 캐시 SET: {test_data}")
    
    # page_obj 변수 초기화
    page_obj = None
    
    # POST 요청 처리 (그룹화, 삭제, 수정 등)
    if request.method == 'POST':
        # 전체 주요 설정 변경 처리
        if 'bulk_important_all' in request.POST:
            is_important = request.POST.get('bulk_important_all') == '1'
            
            # 모든 사용자의 PurchaseGroup의 is_important 상태를 일괄 변경
            all_user_groups = PurchaseGroup.objects.filter(user=request.user, is_active=True)
            all_user_groups.update(is_important=is_important)
            
            messages.success(request, f"모든 퍼블리셔의 주요 설정이 {'설정' if is_important else '해제'}되었습니다.")
            redirect_url = f'/purchase-report/?year={year}&search={search_query}'
            return redirect(redirect_url)
        
        # 주요 퍼블리셔 설정 처리 (수기 항목 저장이 아닌 경우에만 실행)
        if 'save_changes' not in request.POST:
            important_publishers = set()
            for key, value in request.POST.items():
                if key.startswith('is_important_') and value == '1':
                    publisher_code = key.replace('is_important_', '')
                    important_publishers.add(publisher_code)
            
            # 모든 PurchaseGroup의 is_important 상태 업데이트
            all_user_groups = PurchaseGroup.objects.filter(user=request.user, is_active=True)
            for group in all_user_groups:
                is_important = group.member.request_key in important_publishers
                if group.is_important != is_important:
                    group.is_important = is_important
                    group.save()
            
            # 그룹이 없는 멤버를 주요 퍼블리셔로 설정할 때 그룹 생성
            for publisher_code in important_publishers:
                try:
                    member = Member.newspic_objects().get(request_key=publisher_code)
                    existing_group = PurchaseGroup.objects.filter(member_request_key=publisher_code, user=request.user, is_active=True).first()
                    
                    if not existing_group:
                        # 그룹이 없으면 생성
                        PurchaseGroup.objects.create(
                            user=request.user,
                            member_request_key=publisher_code,
                            group_name=f"{member.uname or '미설정'} 그룹",
                            service_name=member.uname or "미설정",
                            company_name=member.uname or "미설정",
                            default_unit_price=Decimal('50'),
                            default_unit_type='percent',
                            is_active=True,
                            is_important=True
                        )
                except Member.DoesNotExist:
                    continue
        
        # 수기 입력 항목 저장 로직
        if 'save_changes' in request.POST:
            # 화면에 표시된 모든 항목의 코드를 가져옵니다.
            submitted_codes = request.POST.getlist('row_codes')
            
            # '주요'로 체크된 항목들의 코드를 집합으로 만듭니다.
            important_codes = set()
            for key, value in request.POST.items():
                if key.startswith('is_important_') and value == '1':
                    code = key.replace('is_important_', '')
                    important_codes.add(code)

            # 서비스명, 거래처명, 단가 등 다른 수기 입력 항목들을 가져옵니다.
            updates = {}
            for key, value in request.POST.items():
                if '_' in key:
                    parts = key.split('_', 2)
                    if len(parts) == 3:
                        prefix = f"{parts[0]}_{parts[1]}"
                        code = parts[2]
                    else:
                        prefix = parts[0]
                        code = parts[1]

                    if prefix in ['service_name', 'company_name', 'unit_price', 'unit_type']:
                        if code not in updates:
                            updates[code] = {}
                        updates[code][prefix] = value
            
            success_count = 0
            error_count = 0
            
            # 화면에 표시된 각 항목에 대해 업데이트를 처리합니다.
            for code in submitted_codes:
                try:
                    member = Member.newspic_objects().get(request_key=code)
                    pg, created = PurchaseGroup.objects.get_or_create(
                        member_request_key=code, 
                        user=request.user,
                        defaults={
                            'group_name': f"{(member.uname or code)} 그룹",
                            'service_name': member.uname or "",
                            'company_name': member.uname or "",
                            'default_unit_price': Decimal('50.00'),
                            'default_unit_type': 'percent',
                            'is_active': True
                        }
                    )

                    data = updates.get(code, {})
                    
                    # 폼에서 받은 데이터로 각 필드를 업데이트합니다.
                    pg.service_name = data.get('service_name', pg.service_name)
                    pg.company_name = data.get('company_name', pg.company_name)
                    unit_price_str = data.get('unit_price', str(pg.default_unit_price))
                    pg.default_unit_price = Decimal(unit_price_str) if unit_price_str and unit_price_str != 'None' else Decimal('0')
                    pg.default_unit_type = data.get('unit_type', pg.default_unit_type)
                    
                    # '주요' 체크박스 상태를 업데이트합니다.
                    pg.is_important = code in important_codes
                    
                    pg.save()
                    success_count += 1

                except Member.DoesNotExist:
                    error_count += 1
                    print(f"Member with request_key {code} not found. Skipping.")
                    continue
                except Exception as e:
                    error_count += 1
                    print(f"Error updating purchase data for {code}: {e}")
            
            if success_count:
                messages.success(request, f"{success_count}개 항목이 성공적으로 업데이트되었습니다.")
            if error_count:
                messages.error(request, f"{error_count}개 항목 업데이트에 실패했습니다.")

            # 처리가 끝난 후, 현재 필터 상태를 유지하며 페이지를 새로고침합니다.
            redirect_url = f'/purchase-report/?year={year}'
            if search_query:
                redirect_url += f'&search={search_query}'
            
            redirect_url += f'&important_only={1 if important_only else 0}'
            
            return redirect(redirect_url)

        if 'bulk-edit-form' in request.POST:
            # 벌크 수정은 현재 모델 구조에서 지원하지 않으므로 기본 메시지만 표시
            messages.info(request, "벌크 수정 기능은 현재 지원되지 않습니다.")
            return redirect(f'/purchase-report/?year={year}&month={month}&search={search_query}')

        if 'delete-group' in request.POST:
            group_code = request.POST.get('group_code')
            try:
                # PurchaseGroup에는 group_code 필드가 없으므로 다른 방식으로 찾기
                # 여기서는 member의 request_key를 사용
                member_key = request.POST.get('member_key')
                if member_key:
                    group = PurchaseGroup.objects.filter(member_request_key=member_key, is_active=True).first()
                    if group:
                        group.is_active = False
                        group.save()
                        messages.success(request, f"그룹 '{group.group_name}'이(가) 비활성화되었습니다.")
                    else:
                        messages.error(request, "삭제하려는 그룹을 찾을 수 없습니다.")
                else:
                    messages.error(request, "그룹 정보가 올바르지 않습니다.")
            except Exception as e:
                messages.error(request, f"그룹 삭제 중 오류가 발생했습니다: {e}")
            return redirect(f'/purchase-report/?year={year}&month={month}&search={search_query}')

        if 'group-form' in request.POST:
            # 그룹화 기능은 현재 모델 구조에서 지원하지 않으므로 기본 메시지만 표시
            messages.info(request, "그룹화 기능은 현재 지원되지 않습니다.")
            return redirect(f'/purchase-report/?year={year}&month={month}&search={search_query}')

        if 'ungroup-selected' in request.POST:
            # 그룹 해제 기능은 현재 모델 구조에서 지원하지 않으므로 기본 메시지만 표시
            messages.info(request, "그룹 해제 기능은 현재 지원되지 않습니다.")
            return redirect(f'/purchase-report/?year={year}&month={month}&search={search_query}')

        # 인라인 수정 처리 - 현재는 PurchaseGroup의 기본 단가만 수정 가능
        for key, value in request.POST.items():
            try:
                parts = key.split('_')
                if len(parts) >= 3:
                    field_type = parts[0]
                    field_name = parts[1]
                    instance_id = parts[2]

                    if field_type == 'PurchaseGroup':
                        # PurchaseGroup 수정 (member의 request_key로 찾기)
                        try:
                            member = Member.newspic_objects().get(request_key=instance_id)
                            group = PurchaseGroup.objects.filter(member_request_key=instance_id, is_active=True).first()
                            if group and hasattr(group, field_name):
                                if field_name in ['default_unit_price']:
                                    try:
                                        value = Decimal(value) if value else Decimal('50')
                                    except:
                                        messages.error(request, f"'{value}'는 유효한 숫자가 아닙니다.")
                                        continue
                                setattr(group, field_name, value)
                                group.save()
                        except Member.DoesNotExist:
                            continue
                    # AdStats 수정은 현재 모델 구조에서 지원하지 않음
            except (IndexError, ValueError):
                continue # 정상적인 폼 데이터가 아닌 경우 무시 (e.g. csrf token)
            except Exception as e:
                messages.error(request, f"업데이트 중 오류 발생: {e}")
                continue
        
        return redirect(f'/purchase-report/?year={year}&month={month}&search={search_query}')

    # GET 요청 처리
    months = list(range(1, 13))
    
    # 검색어가 없고 important_only가 True이면 주요 퍼블리셔만 표시
    if not search_query and important_only:
        # 주요 퍼블리셔만 기본으로 표시
        important_groups = PurchaseGroup.objects.filter(
            user=request.user, 
            is_active=True,
            is_important=True
        ).prefetch_related('ad_units')
        
        # 주요 퍼블리셔의 멤버들
        important_member_keys = {group.member.request_key for group in important_groups}
        
        # 주요 퍼블리셔 데이터 처리 (성능 최적화된 일별 계산)
        # 캐시 키 생성 (더 구체적으로)
        cache_key = f"purchase_data_important_{request.user.id}_{year}_{request.session.session_key}"
        
        # 캐시에서 데이터 확인
        cached_data = cache.get(cache_key)
        if cached_data:
            logger.info(f"캐시 HIT: {cache_key} - 주요 퍼블리셔 데이터 로드")
            purchase_data = cached_data['purchase_data']
            publisher_total = cached_data['publisher_total']
            partners_total = cached_data['partners_total']
            
            # 캐시에서 데이터를 가져왔으면 계산 생략하고 바로 context 생성
            total = {m: publisher_total[m] + partners_total[m] for m in months}
            
            # 월별 증감 계산
            def calculate_monthly_changes(monthly_data):
                changes = {}
                for month in months:
                    current = monthly_data.get(month, Decimal('0'))
                    if month == 1:
                        changes[month] = {'amount': current, 'percent': 100 if current > 0 else 0}
                    else:
                        previous = monthly_data.get(month - 1, Decimal('0'))
                        change_amount = current - previous
                        if previous > 0:
                            change_percent = (change_amount / previous) * 100
                        else:
                            change_percent = 100 if current > 0 else 0
                        changes[month] = {'amount': change_amount, 'percent': change_percent}
                return changes
            
            # 각 퍼블리셔별 증감 계산
            for item in purchase_data:
                item['monthly_changes'] = calculate_monthly_changes(item['monthly_cost'])
            
            # 합계별 증감 계산
            publisher_changes = calculate_monthly_changes(publisher_total)
            partners_changes = calculate_monthly_changes(partners_total)
            total_changes = calculate_monthly_changes(total)
            
            context = {
                'year': year,
                'month': month,
                'months': months,
                'purchase_data': purchase_data,
                'publisher_total': publisher_total,
                'partners_total': partners_total,
                'total': total,
                'publisher_changes': publisher_changes,
                'partners_changes': partners_changes,
                'total_changes': total_changes,
                'publisher_details': [],
                'search_query': search_query,
                'showing_important_only': important_only,
                'important_only': important_only,
                'page_obj': page_obj,
            }
            return render(request, 'purchase.html', context)
        
        # 캐시 MISS인 경우에만 계산 시작
        logger.info(f"캐시 MISS: {cache_key} - 주요 퍼블리셔 데이터 계산 시작")
        purchase_data = []
        publisher_total = {m: Decimal('0') for m in months}
        partners_total = {m: Decimal('0') for m in months}
        
        # 모든 광고 단위 ID 수집
        all_ad_unit_ids = []
        for group in important_groups:
            for ad_unit in group.ad_units.all():
                if ad_unit.is_active:
                    all_ad_unit_ids.append(ad_unit.ad_unit_id)
        
        # 1. AdStats 데이터 일괄 조회 (일별 데이터)
        ad_stats = AdStats.objects.filter(
            user=request.user,
            ad_unit_id__in=all_ad_unit_ids,
            date__year=year
        ).values('date', 'platform', 'ad_unit_id').annotate(
            earnings=Sum('earnings'),
            earnings_usd=Sum('earnings_usd')
        )
        
        # 2. 환율 데이터 일괄 조회
        exchange_rates = ExchangeRate.objects.filter(
            user=request.user, 
            year_month__gte=date(year, 1, 1)
        ).values('year_month', 'usd_to_krw')
        exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}
        
        # 3. stats_map 생성 (세부 데이터와 동일한 방식)
        stats_map = {}
        for stat in ad_stats:
            key = (stat['date'], stat['ad_unit_id'])
            if stat['platform'] == 'adsense':
                month_start = stat['date'].replace(day=1)
                exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
                stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
            elif stat['platform'] == 'admanager':
                stats_map[key] = stat['earnings'] or 0
            elif stat['platform'] == 'adpost':
                stats_map[key] = stat['earnings'] or 0
            else:
                stats_map[key] = stat['earnings'] or 0
        
        # 4. 파워링크(애드포스트) 수익 분배 로직
        # 4-1. 파워링크(애드포스트) 데이터 일괄 조회
        adpost_stats = AdStats.objects.filter(
            platform='adpost',
            ad_unit_id='모바일뉴스픽_컨텐츠',
            date__year=year,
            user=request.user
        ).values('date').annotate(
            earnings=Sum('earnings'),
            clicks=Sum('clicks')
        )
        adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
        
        # 4-2. 전체 파워링크 클릭수 일괄 조회
        total_powerlink_stats = TotalStat.newspic_objects().filter(
            sdate__year=year
        ).values('sdate').annotate(
            total_powerlink=Sum('powerlink_count')
        )
        total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
        
        # 4-3. 퍼블리셔별 파워링크 클릭수 일괄 조회
        publisher_keys = [group.member.request_key for group in important_groups]
        member_powerlink_stats = TotalStat.newspic_objects().filter(
            request_key__in=publisher_keys,
            sdate__year=year
        ).values('request_key', 'sdate').annotate(
            powerlink_count=Sum('powerlink_count'),
            click_count=Sum('click_count')
        )
        member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}
        
        for group in important_groups:
            monthly_cost = {m: Decimal('0') for m in months}
            default_price = group.default_unit_price
            default_type = group.default_unit_type
            ad_units = group.ad_units.filter(is_active=True)
            
            # 해당 연도의 모든 날짜에 대해 일별 계산 (세부 데이터와 동일한 방식)
            for m in months:
                last_day = calendar.monthrange(year, m)[1]
                for d in range(1, last_day + 1):
                    current_date = date(year, m, d)
                    
                    # 1. AdSense/AdManager 일별 수익 계산 (세부 데이터와 동일한 방식)
                    ad_revenue = Decimal('0')
                    for ad_unit in ad_units:
                        stat_key = (current_date, ad_unit.ad_unit_id)
                        ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
                    
                    # 2. AdPost 일별 수익 분배 (세부 데이터와 동일한 방식)
                    adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                    total_powerlink_count = total_powerlink_data.get(current_date, 0)
                    member_data = member_powerlink_data.get((group.member.request_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                    powerlink_count = member_data['powerlink_count']
                    click_count = member_data['click_count']
                    
                    adpost_earnings = adpost_info['earnings']
                    adpost_clicks = adpost_info['clicks']
                    powerlink_revenue = Decimal('0')
                    
                    if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                        unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                    
                    ad_revenue += powerlink_revenue
                    
                    # 3. 매입비용 계산 (세부 데이터와 동일한 방식)
                    rs_rate = default_price or 0
                    rs_type = default_type or 'percent'
                    purchase_cost = Decimal('0')
                    
                    if rs_type == 'percent':
                        purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
                    else:
                        # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                        purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
                    
                    # 4. 월별 매입비용에 누적 (세부 데이터와 동일한 반올림 방식)
                    monthly_cost[m] += int(purchase_cost)
            
            # 합계 반영
            if group.member.level == 50:
                for m in months: publisher_total[m] += monthly_cost[m]
            elif group.member.level == 60:
                for m in months: partners_total[m] += monthly_cost[m]
            
            purchase_data.append({
                'service_name': group.service_name,
                'company_name': group.company_name,
                'unit_price': str(default_price) if default_price is not None else '0',
                'unit_type': default_type,
                'publisher_code': group.member.request_key,
                'monthly_cost': monthly_cost,
                'has_group': True,
                'member': group.member,
                'is_important': group.is_important,
            })
        
        total = {m: publisher_total[m] + partners_total[m] for m in months}
        
        # 월별 증감 계산
        def calculate_monthly_changes(monthly_data):
            changes = {}
            for month in months:
                current = monthly_data.get(month, Decimal('0'))
                if month == 1:
                    # 1월은 이전 월 데이터가 없으므로 0
                    changes[month] = {'amount': current, 'percent': 100 if current > 0 else 0}
                else:
                    previous = monthly_data.get(month - 1, Decimal('0'))
                    change_amount = current - previous
                    if previous > 0:
                        change_percent = (change_amount / previous) * 100
                    else:
                        change_percent = 100 if current > 0 else 0
                    changes[month] = {'amount': change_amount, 'percent': change_percent}
            return changes
        
        # 각 퍼블리셔별 증감 계산
        for item in purchase_data:
            item['monthly_changes'] = calculate_monthly_changes(item['monthly_cost'])
        
        # 합계별 증감 계산
        publisher_changes = calculate_monthly_changes(publisher_total)
        partners_changes = calculate_monthly_changes(partners_total)
        total_changes = calculate_monthly_changes(total)
        
        # 24시간 캐시 저장 (86400초)
        cache.set(cache_key, {
            'purchase_data': purchase_data,
            'publisher_total': publisher_total,
            'partners_total': partners_total
        }, 86400)
        logger.info(f"캐시 SET: {cache_key} - 주요 퍼블리셔 데이터 저장 완료")
        
        context = {
            'year': year,
            'month': month,
            'months': months,
            'purchase_data': purchase_data,
            'publisher_total': publisher_total,
            'partners_total': partners_total,
            'total': total,
            'publisher_changes': publisher_changes,
            'partners_changes': partners_changes,
            'total_changes': total_changes,
            'publisher_details': [], # 세부 데이터 로직은 일단 비워둠
            'search_query': search_query,
            'showing_important_only': important_only,
            'important_only': important_only,
            'page_obj': page_obj,  # 페이징 객체 추가
        }
        return render(request, 'purchase.html', context)
    
    # 검색어가 있을 때만 데이터 조회
    # 1. 검색 조건에 맞는 PurchaseGroup 데이터 가져오기 (현재 사용자만)
    all_groups = PurchaseGroup.objects.filter(
        user=request.user, 
        is_active=True
    ).prefetch_related('ad_units')
    
    # 2. 검색 조건에 맞는 Member 데이터 가져오기 (최적화된 방법)
    # 먼저 필요한 request_key들만 수집
    needed_request_keys = set()
    
    # 검색어가 있으면 검색 결과를 우선적으로 추가
    if search_query:
        # 검색어로 필터링된 Member 조회
        search_terms = [term.strip() for term in search_query.split(',') if term.strip()]
        member_query = Q()
        for term in search_terms:
            member_query |= (
                Q(request_key__icontains=term) |
                Q(uname__icontains=term)
            )
        
        # 검색 결과에서 request_key만 가져오기 (level 50, 60만)
        search_members = Member.newspic_objects().filter(member_query, level__in=[50, 60]).values_list('request_key', flat=True)
        needed_request_keys.update(search_members)
        
        # 검색어로 그룹도 필터링
        group_query = Q()
        for term in search_terms:
            group_query |= (
                Q(member_request_key__icontains=term) |
                Q(company_name__icontains=term) |
                Q(service_name__icontains=term)
            )
        all_groups = all_groups.filter(group_query)
    
    # 그룹에 속한 멤버들의 request_key 수집 (검색어가 없거나 추가로 필요한 경우)
    for group in all_groups:
        needed_request_keys.add(group.member_request_key)
    
    # important_only가 True이고 검색어가 없으면 주요 항목만 필터링
    if important_only and not search_query:
        all_groups = all_groups.filter(is_important=True)
        # 주요 퍼블리셔의 멤버들만 유지
        important_member_keys = {group.member.request_key for group in all_groups}
        needed_request_keys = needed_request_keys.intersection(important_member_keys)
    elif not search_query and not important_only:
        # 검색어가 없고 주요항목만이 False이면 level 50, 60인 Member만 표시
        # 그룹에 속하지 않은 Member들도 포함하여 표시
        all_members_queryset = Member.newspic_objects().filter(level__in=[50, 60]).order_by('-level', 'uname')
        # 페이징을 위해 처음 1000개만 가져오기 (성능 고려)
        all_members_queryset = all_members_queryset[:1000]
        needed_request_keys.update(all_members_queryset.values_list('request_key', flat=True))
    
    # 필요한 Member만 조회 (level 50, 60만)
    if needed_request_keys:
        # request_key 리스트를 청크로 나누어 처리 (메모리 효율성)
        request_key_list = list(needed_request_keys)
        members_queryset = Member.newspic_objects().filter(
            request_key__in=request_key_list,
            level__in=[50, 60]
        ).order_by('-level', 'uname')
    else:
        members_queryset = Member.newspic_objects().none()
    
    # 검색어가 있으면 페이징 없이 모든 결과 표시, 없으면 페이징 적용
    if search_query:
        # 검색 결과는 페이징 없이 모든 결과 표시
        all_members = list(members_queryset)
        page_obj = None
    else:
        # 검색어가 없으면 페이징 적용 (페이지당 100개씩)
        page = request.GET.get('page', 1)
        try:
            page = int(page)
        except ValueError:
            page = 1
        
        from django.core.paginator import Paginator
        paginator = Paginator(members_queryset, 100)
        all_members = paginator.get_page(page)
        page_obj = all_members
    
    # 딕셔너리 생성 최적화
    all_members_map = {member.request_key: member for member in all_members}
    
    # 3. 월별 합계 초기화
    # 캐시 키 생성 (더 구체적으로)
    cache_key = f"purchase_data_all_{request.user.id}_{year}_{search_query}_{important_only}_{request.session.session_key}"
    
    # 캐시에서 데이터 확인
    cached_data = cache.get(cache_key)
    if cached_data:
        logger.info(f"캐시 HIT: {cache_key} - 전체 퍼블리셔 데이터 로드")
        purchase_data = cached_data['purchase_data']
        publisher_total = cached_data['publisher_total']
        partners_total = cached_data['partners_total']
        
        # 캐시에서 데이터를 가져왔으면 계산 생략하고 바로 context 생성
        total = {m: publisher_total[m] + partners_total[m] for m in months}
        
        # 월별 증감 계산
        def calculate_monthly_changes(monthly_data):
            changes = {}
            for month in months:
                current = monthly_data.get(month, Decimal('0'))
                if month == 1:
                    changes[month] = {'amount': current, 'percent': 100 if current > 0 else 0}
                else:
                    previous = monthly_data.get(month - 1, Decimal('0'))
                    change_amount = current - previous
                    if previous > 0:
                        change_percent = (change_amount / previous) * 100
                    else:
                        change_percent = 100 if current > 0 else 0
                    changes[month] = {'amount': change_amount, 'percent': change_percent}
            return changes
        
        # 각 퍼블리셔별 증감 계산
        for item in purchase_data:
            item['monthly_changes'] = calculate_monthly_changes(item['monthly_cost'])
        
        # 합계별 증감 계산
        publisher_changes = calculate_monthly_changes(publisher_total)
        partners_changes = calculate_monthly_changes(partners_total)
        total_changes = calculate_monthly_changes(total)
        
        context = {
            'year': year,
            'month': month,
            'months': months,
            'purchase_data': purchase_data,
            'publisher_total': publisher_total,
            'partners_total': partners_total,
            'total': total,
            'publisher_changes': publisher_changes,
            'partners_changes': partners_changes,
            'total_changes': total_changes,
            'publisher_details': [],
            'search_query': search_query,
            'important_only': important_only,
            'all_members': all_members,
            'page_obj': all_members,
        }
        return render(request, 'purchase.html', context)
    
    # 캐시 MISS인 경우에만 계산 시작
    logger.info(f"캐시 MISS: {cache_key} - 전체 퍼블리셔 데이터 계산 시작")
    publisher_total = {m: Decimal('0') for m in months}
    partners_total = {m: Decimal('0') for m in months}
    
    purchase_data = []
    
    # 4. 그룹별 데이터 처리 - 각 멤버별로 개별 행 표시
    
    # 모든 광고 단위 ID 수집 (메모리에서 처리)
    all_ad_unit_ids = []
    for group in all_groups:
        for ad_unit in group.ad_units.all():
            if ad_unit.is_active:
                all_ad_unit_ids.append(ad_unit.ad_unit_id)
    
    # AdStats 데이터 일괄 조회 (세부 데이터와 동일한 방식)
    ad_stats = AdStats.objects.filter(
        user=request.user,
        ad_unit_id__in=all_ad_unit_ids,
        date__year=year
    ).values('date', 'platform', 'ad_unit_id').annotate(
        earnings=Sum('earnings'),
        earnings_usd=Sum('earnings_usd')
    )
    
    # 환율 데이터 일괄 조회 (세부 데이터와 동일한 방식)
    exchange_rates = ExchangeRate.objects.filter(
        user=request.user, 
        year_month__gte=date(year, 1, 1)
    ).values('year_month', 'usd_to_krw')
    exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}
    
    # stats_map 생성 (세부 데이터와 동일한 방식)
    stats_map = {}
    for stat in ad_stats:
        key = (stat['date'], stat['ad_unit_id'])
        if stat['platform'] == 'adsense':
            month_start = stat['date'].replace(day=1)
            exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
            stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
        elif stat['platform'] == 'admanager':
            stats_map[key] = stat['earnings'] or 0
        elif stat['platform'] == 'adpost':
            stats_map[key] = stat['earnings'] or 0
        else:
            stats_map[key] = stat['earnings'] or 0
    
    # 애드포스트(파워링크) 데이터 일괄 조회
    adpost_data = {}
    adpost_stats = AdStats.objects.filter(
        platform='adpost',
        ad_unit_id='모바일뉴스픽_컨텐츠',
        date__year=year,
        user=request.user
    ).values('date').annotate(
        earnings=Sum('earnings'),
        clicks=Sum('clicks')
    )
    adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
    
    # 전체 파워링크 클릭수 일괄 조회
    total_powerlink_data = {}
    total_powerlink_stats = TotalStat.newspic_objects().filter(
        sdate__year=year
    ).values('sdate').annotate(
        total_powerlink=Sum('powerlink_count')
    )
    total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
    
    # 퍼블리셔별 파워링크 클릭수 일괄 조회
    member_powerlink_data = {}
    publisher_keys = [group.member.request_key for group in all_groups]
    member_powerlink_stats = TotalStat.newspic_objects().filter(
        request_key__in=publisher_keys, sdate__year=year
    ).values('request_key', 'sdate').annotate(
        powerlink_count=Sum('powerlink_count'),
        click_count=Sum('click_count')
    )
    member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}
    
    for group in all_groups:
        monthly_cost = {m: Decimal('0') for m in months}
        default_price = group.default_unit_price
        default_type = group.default_unit_type
        ad_units = group.ad_units.filter(is_active=True)
        
        # 해당 연도의 모든 날짜에 대해 일별 계산 (세부 데이터와 동일한 방식)
        for m in months:
            last_day = calendar.monthrange(year, m)[1]
            for d in range(1, last_day + 1):
                current_date = date(year, m, d)
                
                # 1. AdSense/AdManager 일별 수익 계산 (세부 데이터와 동일한 방식)
                ad_revenue = Decimal('0')
                for ad_unit in ad_units:
                    stat_key = (current_date, ad_unit.ad_unit_id)
                    ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
                
                # 2. AdPost 일별 수익 분배 (세부 데이터와 동일한 방식)
                adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                total_powerlink_count = total_powerlink_data.get(current_date, 0)
                member_data = member_powerlink_data.get((group.member.request_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                powerlink_count = member_data['powerlink_count']
                click_count = member_data['click_count']
                adpost_earnings = adpost_info['earnings']
                adpost_clicks = adpost_info['clicks']
                powerlink_revenue = Decimal('0')
                if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                    unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                ad_revenue += powerlink_revenue
                
                # 3. 매입비용 계산 (세부 데이터와 동일한 방식)
                rs_rate = default_price or 0
                rs_type = default_type or 'percent'
                purchase_cost = Decimal('0')
                
                if rs_type == 'percent':
                    purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
                else:
                    # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                    purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
                
                # 4. 월별 매입비용에 누적 (세부 데이터와 동일한 반올림 방식)
                monthly_cost[m] += int(purchase_cost)
        
        # 합계 반영
        if group.member.level == 50:
            for m in months: publisher_total[m] += monthly_cost[m]
        elif group.member.level == 60:
            for m in months: partners_total[m] += monthly_cost[m]
        
        purchase_data.append({
            'service_name': group.service_name,
            'company_name': group.company_name,
            'unit_price': str(default_price) if default_price is not None else '0',
            'unit_type': default_type,
            'publisher_code': group.member_request_key,
            'monthly_cost': monthly_cost,
            'has_group': True,
            'member': group.member,
            'is_important': group.is_important,
        })
    
    # 5. 그룹에 속하지 않은 멤버 처리
    grouped_member_keys = {group.member_request_key for group in all_groups}
    
    # 검색 결과에 포함된 그룹에 속하지 않은 멤버들도 처리
    ungrouped_members = []
    for request_key, member in all_members_map.items():
        if request_key in grouped_member_keys:
            continue
        
        # 그룹에 속하지 않은 멤버의 광고 단위 ID 수집
        member_ad_units = PurchaseGroupAdUnit.objects.filter(
            purchase_group__member_request_key=request_key,
            purchase_group__user=request.user,
            purchase_group__is_active=True,
            is_active=True
        )
        
        # 그룹이 없는 멤버가 주요 퍼블리셔로 설정되어 있는지 확인
        is_important = False
        inactive_group = PurchaseGroup.objects.filter(
            member_request_key=request_key, 
            user=request.user, 
            is_active=False
        ).first()
        if inactive_group and inactive_group.is_important:
            is_important = True
        
        ungrouped_members.append({
            'request_key': request_key,
            'member': member,
            'ad_units': member_ad_units,
            'is_important': is_important
        })
    
    # 그룹에 속하지 않은 멤버의 AdStats 데이터 일괄 조회
    ungrouped_ad_unit_ids = []
    for ungrouped_member in ungrouped_members:
        ungrouped_ad_unit_ids.extend([ad_unit.ad_unit_id for ad_unit in ungrouped_member['ad_units']])
    
    ungrouped_ad_stats = {}
    if ungrouped_ad_unit_ids:
        ungrouped_ad_stats_data = AdStats.objects.filter(
            user=request.user,
            ad_unit_id__in=ungrouped_ad_unit_ids
        ).values('date', 'platform', 'ad_unit_id', 'earnings', 'earnings_usd')
        
        for stat in ungrouped_ad_stats_data:
            key = (stat['date'], stat['ad_unit_id'])
            if key not in ungrouped_ad_stats:
                ungrouped_ad_stats[key] = {'earnings': 0, 'earnings_usd': 0}
            ungrouped_ad_stats[key]['earnings'] += stat['earnings'] or 0
            ungrouped_ad_stats[key]['earnings_usd'] += stat['earnings_usd'] or 0
    
    for ungrouped_member in ungrouped_members:
        request_key = ungrouped_member['request_key']
        member = ungrouped_member['member']
        member_ad_units = ungrouped_member['ad_units']
        is_important = ungrouped_member['is_important']
        
        monthly_cost = {m: Decimal('0') for m in months}
        default_unit_price = Decimal('50')
        default_unit_type = 'percent'
        
        # AdStats에 데이터가 있는 멤버만 실제 매입 비용 계산
        all_member_aliases = AdStats.objects.filter(
            platform='member', 
            user=request.user
        ).values_list('alias', flat=True).distinct()
        
        if request_key in all_member_aliases:
            # 해당 연도의 모든 날짜에 대해 일별 계산 (세부 데이터와 동일한 방식)
            for m in months:
                last_day = calendar.monthrange(year, m)[1]
                for d in range(1, last_day + 1):
                    current_date = date(year, m, d)
                    
                    # 1. AdSense/AdManager 일별 수익 계산 (세부 데이터와 동일한 방식)
                    ad_revenue = Decimal('0')
                    for ad_unit in member_ad_units:
                        stat_key = (current_date, ad_unit.ad_unit_id)
                        if stat_key in ungrouped_ad_stats:
                            stat_data = ungrouped_ad_stats[stat_key]
                            if ad_unit.platform == 'adsense':
                                # AdSense는 USD를 KRW로 변환 (세부 데이터와 동일한 방식)
                                month_start = date(year, m, 1)
                                usd_to_krw = exchange_rates.get(month_start, Decimal('1370.00'))
                                earnings_usd = Decimal(str(stat_data['earnings_usd'] or 0))
                                ad_revenue += earnings_usd * usd_to_krw
                            elif ad_unit.platform == 'admanager':
                                # AdManager는 KRW 그대로
                                earnings = Decimal(str(stat_data['earnings'] or 0))
                                ad_revenue += earnings
                    
                    # 2. AdPost 일별 수익 분배 (세부 데이터와 동일한 방식)
                    adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                    total_powerlink_count = total_powerlink_data.get(current_date, 0)
                    member_data = member_powerlink_data.get((request_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                    powerlink_count = member_data['powerlink_count']
                    click_count = member_data['click_count']
                    adpost_earnings = adpost_info['earnings']
                    adpost_clicks = adpost_info['clicks']
                    powerlink_revenue = Decimal('0')
                    if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                        unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                    ad_revenue += powerlink_revenue
                    
                    # 3. RS단가 적용하여 일별 매입비용 계산
                    month_start = date(year, m, 1)
                    monthly_price = PurchasePrice.objects.filter(
                        user=request.user,
                        request_key=request_key,
                        year_month=month_start
                    ).first()
                    
                    if monthly_price:
                        unit_price = monthly_price.unit_price
                        unit_type = monthly_price.unit_type
                    else:
                        unit_price = default_unit_price
                        unit_type = default_unit_type
                    
                    purchase_cost = Decimal('0')
                    if unit_price is not None:
                        if unit_type == 'percent':
                            purchase_cost = ad_revenue * (Decimal(str(unit_price)) / Decimal('100'))
                        else:
                            # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                            purchase_cost = Decimal(str(click_count)) * Decimal(str(unit_price))
                    
                    # 4. 월별 매입비용에 누적
                    monthly_cost[m] += int(purchase_cost)
        
        # 매입 비용이 있는 멤버만 합계에 포함
        if any(monthly_cost.values()):
            if member.level == 50: # 퍼블리셔
                for m in months: publisher_total[m] += monthly_cost[m]
            elif member.level in [60, 61, 65]: # 파트너스
                for m in months: partners_total[m] += monthly_cost[m]
        
        purchase_data.append({
            'service_name': f'{member.uname or "미설정"}',
            'company_name': f'{member.uname or "미설정"}',
            'unit_price': str(default_unit_price),
            'unit_type': default_unit_type,
            'publisher_code': request_key,
            'monthly_cost': monthly_cost,
            'has_group': False,
            'member': member,
            'is_important': is_important,
        })
    
    # 6. 총 합계 및 컨텍스트
    total = {m: publisher_total[m] + partners_total[m] for m in months}
    
    # 월별 증감 계산
    def calculate_monthly_changes(monthly_data):
        changes = {}
        for month in months:
            current = monthly_data.get(month, Decimal('0'))
            if month == 1:
                # 1월은 이전 월 데이터가 없으므로 0
                changes[month] = {'amount': current, 'percent': 100 if current > 0 else 0}
            else:
                previous = monthly_data.get(month - 1, Decimal('0'))
                change_amount = current - previous
                if previous > 0:
                    change_percent = (change_amount / previous) * 100
                else:
                    change_percent = 100 if current > 0 else 0
                changes[month] = {'amount': change_amount, 'percent': change_percent}
        return changes
    
    # 각 퍼블리셔별 증감 계산
    for item in purchase_data:
        item['monthly_changes'] = calculate_monthly_changes(item['monthly_cost'])
    
    # 합계별 증감 계산
    publisher_changes = calculate_monthly_changes(publisher_total)
    partners_changes = calculate_monthly_changes(partners_total)
    total_changes = calculate_monthly_changes(total)
    
    # 24시간 캐시 저장 (86400초)
    cache.set(cache_key, {
        'purchase_data': purchase_data,
        'publisher_total': publisher_total,
        'partners_total': partners_total
    }, 86400)
    logger.info(f"캐시 SET: {cache_key} - 전체 퍼블리셔 데이터 저장 완료")
    
    context = {
        'year': year,
        'month': month,
        'months': months,
        'purchase_data': purchase_data,
        'publisher_total': publisher_total,
        'partners_total': partners_total,
        'total': total,
        'publisher_changes': publisher_changes,
        'partners_changes': partners_changes,
        'total_changes': total_changes,
        'publisher_details': [], # 세부 데이터 로직은 일단 비워둠
        'search_query': search_query,
        'important_only': important_only,
        # 페이징 정보 추가
        'all_members': all_members,
        'page_obj': all_members,
    }

    return render(request, 'purchase.html', context)

@login_required
def publisher_detail_data_api(request):
    """
    퍼블리셔 세부 데이터 API: AJAX 요청으로 퍼블리셔별 세부 데이터를 JSON으로 반환
    (purchase.html 하단 세부 데이터용)
    """
    try:
        start_date = datetime.strptime(request.GET.get('start_date', ''), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.GET.get('end_date', ''), '%Y-%m-%d').date()
        
        publisher_keys_str = request.GET.get('publishers', '')
        if not publisher_keys_str:
            return JsonResponse({'success': False, 'error': 'No publishers specified.'}, status=400)
        publisher_keys = publisher_keys_str.split(',')

        groups = PurchaseGroup.objects.filter(
            user=request.user,
            is_active=True,
            member_request_key__in=publisher_keys
        )

        publisher_info_map = {
            group.member_request_key: {
                'label': f"{group.company_name} ({group.member.uname if group.member else '미설정'}) ({group.member_request_key})",
                'key': group.member_request_key,
                'rs_rate': group.default_unit_price or 0,
                'rs_type': group.default_unit_type or 'percent',
                'ad_units': list(group.ad_units.filter(is_active=True).values('platform', 'ad_unit_id'))
            } for group in groups
        }
        
        date_list = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # --- 데이터 일괄 조회 ---
        all_ad_units = [unit['ad_unit_id'] for key in publisher_keys for unit in publisher_info_map.get(key, {}).get('ad_units', [])]
        
        ad_stats = AdStats.objects.filter(
            user=request.user,
            date__range=[start_date, end_date],
            ad_unit_id__in=all_ad_units
        ).values('date', 'platform', 'ad_unit_id').annotate(
            earnings=Sum('earnings'),
            earnings_usd=Sum('earnings_usd')
        )

        exchange_rates = ExchangeRate.objects.filter(user=request.user, year_month__gte=start_date.replace(day=1)).values('year_month', 'usd_to_krw')
        exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}

        stats_map = {}
        for stat in ad_stats:
            key = (stat['date'], stat['ad_unit_id'])
            if stat['platform'] == 'adsense':
                month_start = stat['date'].replace(day=1)
                exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
                stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
            elif stat['platform'] == 'admanager':
                # 애드매니저(ADX)는 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
            elif stat['platform'] == 'adpost':
                # 애드포스트는 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
            else:
                # 기타 플랫폼은 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
        
        # --- 파워링크(애드포스트) 수익 분배 로직 추가 ---
        # 1. 파워링크(애드포스트) 데이터 일괄 조회
        adpost_data = {}
        adpost_stats = AdStats.objects.filter(
            platform='adpost',
            ad_unit_id='모바일뉴스픽_컨텐츠',
            date__range=[start_date, end_date],
            user=request.user
        ).values('date').annotate(
            earnings=Sum('earnings'),
            clicks=Sum('clicks')
        )
        adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
        # 2. 전체 파워링크 클릭수 일괄 조회
        total_powerlink_data = {}
        total_powerlink_stats = TotalStat.newspic_objects().filter(
            sdate__range=[start_date, end_date]
        ).values('sdate').annotate(
            total_powerlink=Sum('powerlink_count')
        )
        total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
        # 3. 퍼블리셔별 파워링크 클릭수 일괄 조회
        member_powerlink_data = {}
        publisher_keys = [group.member_request_key for group in groups]
        member_powerlink_stats = TotalStat.newspic_objects().filter(
            request_key__in=publisher_keys, sdate__range=[start_date, end_date]
        ).values('request_key', 'sdate').annotate(
            powerlink_count=Sum('powerlink_count'),
            click_count=Sum('click_count')
        )
        member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}

        # --- 퍼블리셔별 데이터 처리 ---
        all_publishers_detail_data = {}
        for publisher_key in publisher_keys:
            pub_info = publisher_info_map.get(publisher_key)
            if not pub_info:
                continue

            detail_data = []
            detail_totals = {'ad_revenue': 0, 'purchase_cost': 0}
            for current_date in date_list:
                ad_revenue = Decimal('0')
                # 1. 애드센스/애드매니저 광고수익 합산
                for ad_unit in pub_info['ad_units']:
                    stat_key = (current_date, ad_unit['ad_unit_id'])
                    ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
                # 2. 파워링크(애드포스트) 수익 분배
                adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                total_powerlink_count = total_powerlink_data.get(current_date, 0)
                member_data = member_powerlink_data.get((publisher_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                powerlink_count = member_data['powerlink_count']
                click_count = member_data['click_count']
                adpost_earnings = adpost_info['earnings']
                adpost_clicks = adpost_info['clicks']
                powerlink_revenue = Decimal('0')
                if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                    unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                # 매입비용 계산
                rs_rate = pub_info['rs_rate']
                rs_type = pub_info['rs_type']
                purchase_cost = Decimal('0')
                if rs_type == 'percent':
                    purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
                else:
                    # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                    purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
                detail_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'ad_revenue': int(click_count) if rs_type != 'percent' else int(ad_revenue),
                    'ad_revenue_formatted': f"{int(click_count):,}" if rs_type != 'percent' else f"{int(ad_revenue):,}",
                    'rs_rate_display': f"{rs_rate}%" if rs_type == 'percent' else str(rs_rate),
                    'purchase_cost': int(purchase_cost),
                    'purchase_cost_formatted': f"{int(purchase_cost):,}",
                })
                detail_totals['ad_revenue'] += int(click_count) if rs_type != 'percent' else int(ad_revenue)
                detail_totals['purchase_cost'] += int(purchase_cost)
            all_publishers_detail_data[publisher_key] = {
                'label': pub_info['label'],
                'detail_data': detail_data,
                'detail_totals': detail_totals,
                'column_name': '유효페이지뷰' if rs_type != 'percent' else '광고수익',
            }

        return JsonResponse({
            'success': True,
            'publishers': list(all_publishers_detail_data.values())
        })
        
    except Exception as e:
        logger.error(f"퍼블리셔 세부 데이터 API 오류: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def download_publisher_detail_excel(request):
    """
    퍼블리셔 세부 데이터를 엑셀 파일로 다운로드
    """
    try:
        start_date = datetime.strptime(request.GET.get('start_date', ''), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.GET.get('end_date', ''), '%Y-%m-%d').date()
        publisher_keys_str = request.GET.get('publishers', '')
        publisher_keys = publisher_keys_str.split(',') if publisher_keys_str else []

        if not publisher_keys:
            return HttpResponse("조회할 퍼블리셔가 없습니다.", status=400)

        logger.info(f"엑셀 다운로드 시작: 퍼블리셔 {len(publisher_keys)}개, 기간 {start_date} ~ {end_date}")

        # 데이터 조회 로직 (API와 거의 동일)
        groups = PurchaseGroup.objects.filter(
            user=request.user, is_active=True, member_request_key__in=publisher_keys
        )
        
        if not groups.exists():
            logger.warning("활성화된 PurchaseGroup이 없습니다.")
            return HttpResponse("활성화된 퍼블리셔 그룹이 없습니다.", status=400)
        
        logger.info(f"활성화된 그룹 수: {groups.count()}")
        
        publisher_info_map = {
            group.member_request_key: {
                'label': f"{group.company_name} ({group.member.uname if group.member else '미설정'}) ({group.member_request_key})",
                'key': group.member_request_key,
                'rs_rate': group.default_unit_price or 0,
                'rs_type': group.default_unit_type or 'percent',
                'ad_units': list(group.ad_units.filter(is_active=True).values('platform', 'ad_unit_id'))
            } for group in groups
        }
        
        date_list = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        all_ad_units = [unit['ad_unit_id'] for key in publisher_keys for unit in publisher_info_map.get(key, {}).get('ad_units', [])]
        
        ad_stats = AdStats.objects.filter(
            user=request.user,
            date__range=[start_date, end_date],
            ad_unit_id__in=all_ad_units
        ).values('date', 'platform', 'ad_unit_id').annotate(
            earnings=Sum('earnings'),
            earnings_usd=Sum('earnings_usd')
        )

        exchange_rates = ExchangeRate.objects.filter(user=request.user, year_month__gte=start_date.replace(day=1)).values('year_month', 'usd_to_krw')
        exchange_rate_map = {rate['year_month']: rate['usd_to_krw'] for rate in exchange_rates}

        stats_map = {}
        for stat in ad_stats:
            key = (stat['date'], stat['ad_unit_id'])
            if stat['platform'] == 'adsense':
                month_start = stat['date'].replace(day=1)
                exchange_rate = exchange_rate_map.get(month_start, Decimal('1370.00'))
                stats_map[key] = (stat['earnings_usd'] or 0) * float(exchange_rate)
            elif stat['platform'] == 'admanager':
                # 애드매니저(ADX)는 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
            elif stat['platform'] == 'adpost':
                # 애드포스트는 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
            else:
                # 기타 플랫폼은 KRW 그대로
                stats_map[key] = stat['earnings'] or 0
        
        # --- 파워링크(애드포스트) 수익 분배 로직 추가 ---
        # 1. 파워링크(애드포스트) 데이터 일괄 조회
        adpost_data = {}
        adpost_stats = AdStats.objects.filter(
            platform='adpost',
            ad_unit_id='모바일뉴스픽_컨텐츠',
            date__range=[start_date, end_date],
            user=request.user
        ).values('date').annotate(
            earnings=Sum('earnings'),
            clicks=Sum('clicks')
        )
        adpost_data = {stat['date']: {'earnings': stat['earnings'] or 0, 'clicks': stat['clicks'] or 0} for stat in adpost_stats}
        
        # 2. 전체 파워링크 클릭수 일괄 조회
        total_powerlink_data = {}
        total_powerlink_stats = TotalStat.newspic_objects().filter(
            sdate__range=[start_date, end_date]
        ).values('sdate').annotate(
            total_powerlink=Sum('powerlink_count')
        )
        total_powerlink_data = {stat['sdate']: stat['total_powerlink'] or 0 for stat in total_powerlink_stats}
        
        # 3. 퍼블리셔별 파워링크 클릭수 일괄 조회
        member_powerlink_data = {}
        publisher_keys_for_powerlink = [group.member_request_key for group in groups]
        member_powerlink_stats = TotalStat.newspic_objects().filter(
            request_key__in=publisher_keys_for_powerlink, sdate__range=[start_date, end_date]
        ).values('request_key', 'sdate').annotate(
            powerlink_count=Sum('powerlink_count'),
            click_count=Sum('click_count')
        )
        member_powerlink_data = {(stat['request_key'], stat['sdate']): {'powerlink_count': stat['powerlink_count'] or 0, 'click_count': stat['click_count'] or 0} for stat in member_powerlink_stats}
        
        all_publishers_detail_data = {}
        for publisher_key in publisher_keys:
            pub_info = publisher_info_map.get(publisher_key)
            if not pub_info:
                continue

            detail_data = []
            detail_totals = {'ad_revenue': 0, 'purchase_cost': 0}
            for current_date in date_list:
                ad_revenue = Decimal('0')
                # 1. 애드센스/애드매니저 광고수익 합산
                for ad_unit in pub_info['ad_units']:
                    stat_key = (current_date, ad_unit['ad_unit_id'])
                    ad_revenue += Decimal(str(stats_map.get(stat_key, 0)))
                # 2. 파워링크(애드포스트) 수익 분배
                adpost_info = adpost_data.get(current_date, {'earnings': 0, 'clicks': 0})
                total_powerlink_count = total_powerlink_data.get(current_date, 0)
                member_data = member_powerlink_data.get((publisher_key, current_date), {'powerlink_count': 0, 'click_count': 0})
                powerlink_count = member_data['powerlink_count']
                click_count = member_data['click_count']
                adpost_earnings = adpost_info['earnings']
                adpost_clicks = adpost_info['clicks']
                powerlink_revenue = Decimal('0')
                if adpost_clicks > 0 and total_powerlink_count > 0 and powerlink_count > 0:
                    unit_price = (Decimal(str(adpost_earnings)) / Decimal(str(adpost_clicks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    powerlink_revenue = (unit_price * Decimal(str(powerlink_count)) * Decimal('0.595') * (Decimal(str(adpost_clicks)) / Decimal(str(total_powerlink_count)))).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                ad_revenue += powerlink_revenue
                # 매입비용 계산
                rs_rate = pub_info['rs_rate']
                rs_type = pub_info['rs_type']
                purchase_cost = Decimal('0')
                if rs_type == 'percent':
                    purchase_cost = ad_revenue * (Decimal(str(rs_rate)) / Decimal('100'))
                else:
                    # RS단가가 percent가 아닌 경우 tbTotalStat의 click_count에 RS단가를 곱함
                    purchase_cost = Decimal(str(click_count)) * Decimal(str(rs_rate))
                detail_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'ad_revenue': int(click_count) if rs_type != 'percent' else int(ad_revenue),
                    'ad_revenue_formatted': f"{int(click_count):,}" if rs_type != 'percent' else f"{int(ad_revenue):,}",
                    'rs_rate_display': f"{rs_rate}%" if rs_type == 'percent' else str(rs_rate),
                    'purchase_cost': int(purchase_cost),
                    'purchase_cost_formatted': f"{int(purchase_cost):,}",
                })
                detail_totals['ad_revenue'] += int(click_count) if rs_type != 'percent' else int(ad_revenue)
                detail_totals['purchase_cost'] += int(purchase_cost)
            all_publishers_detail_data[publisher_key] = {
                'label': pub_info['label'],
                'detail_data': detail_data,
                'detail_totals': detail_totals,
            }

        # --- 엑셀 생성 ---
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active) # 기본 시트 제거

            # 스타일 정의
            header_font = Font(bold=True, name='맑은 고딕')
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            logger.info(f"엑셀 워크북 생성 완료, 퍼블리셔 수: {len(all_publishers_detail_data)}")
        except Exception as e:
            logger.error(f"엑셀 워크북 생성 실패: {str(e)}")
            raise Exception(f"엑셀 워크북 생성 실패: {str(e)}")

        for pub_key, pub_data in all_publishers_detail_data.items():
            try:
                # 시트명에서 특수문자 제거 및 길이 제한
                safe_title = "".join(c for c in pub_data['label'] if c.isalnum() or c in (' ', '-', '_'))[:31]
                if not safe_title.strip():
                    safe_title = f"Publisher_{pub_key[:20]}"
                
                ws = wb.create_sheet(title=safe_title)
                logger.info(f"시트 생성: {safe_title}")

                # 헤더
                headers = ['일자', '광고수익/유효PV', 'RS율', '매입비용']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # 데이터
                for row_idx, row_data in enumerate(pub_data['detail_data'], 2):
                    try:
                        # 날짜
                        cell = ws.cell(row=row_idx, column=1, value=row_data['date'])
                        cell.border = thin_border
                        cell.alignment = center_align
                        
                        # 광고수익
                        cell = ws.cell(row=row_idx, column=2, value=row_data['ad_revenue'])
                        cell.number_format = '#,##0'
                        cell.border = thin_border
                        cell.alignment = right_align
                        
                        # RS율
                        cell = ws.cell(row=row_idx, column=3, value=row_data['rs_rate_display'])
                        cell.border = thin_border
                        cell.alignment = center_align
                        
                        # 매입비용
                        cell = ws.cell(row=row_idx, column=4, value=row_data['purchase_cost'])
                        cell.number_format = '#,##0'
                        cell.border = thin_border
                        cell.alignment = right_align
                    except Exception as e:
                        logger.error(f"행 데이터 처리 실패 (시트: {safe_title}, 행: {row_idx}): {str(e)}")
                        continue

                # 합계
                total_row_idx = len(pub_data['detail_data']) + 2
                # 합계 라벨
                cell = ws.cell(row=total_row_idx, column=1, value='합계')
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
                cell.fill = header_fill
                
                # 광고수익 합계
                cell = ws.cell(row=total_row_idx, column=2, value=pub_data['detail_totals']['ad_revenue'])
                cell.number_format = '#,##0'
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = right_align
                cell.fill = header_fill
                
                # RS율 합계 (빈칸)
                cell = ws.cell(row=total_row_idx, column=3, value='')
                cell.border = thin_border
                cell.fill = header_fill
                
                # 매입비용 합계
                cell = ws.cell(row=total_row_idx, column=4, value=pub_data['detail_totals']['purchase_cost'])
                cell.number_format = '#,##0'
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = right_align
                cell.fill = header_fill

                # 컬럼 너비 조정
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 15
                
                logger.info(f"시트 {safe_title} 완성")
                
            except Exception as e:
                logger.error(f"시트 {pub_key} 생성 실패: {str(e)}")
                continue

        # --- HttpResponse로 반환 ---
        try:
            filename = f"publisher_detail_{start_date}_to_{end_date}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            logger.info(f"엑셀 파일 저장 시작: {filename}")
            wb.save(response)
            logger.info(f"엑셀 파일 저장 완료: {filename}")
            
            return response
            
        except Exception as e:
            logger.error(f"엑셀 파일 저장 실패: {str(e)}")
            raise Exception(f"엑셀 파일 저장 실패: {str(e)}")

    except Exception as e:
        logger.error(f"엑셀 다운로드 오류: {str(e)}", exc_info=True)
        return HttpResponse("엑셀 파일을 생성하는 중 오류가 발생했습니다.", status=500) 

@login_required
def member_search_api(request):
    """Member 검색 API - AJAX용"""
    try:
        search_query = request.GET.get('search', '').strip()
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 50))  # 한 번에 가져올 개수
        
        if not search_query:
            return JsonResponse({'members': [], 'has_more': False})
        
        # 검색어 처리
        search_terms = [term.strip() for term in search_query.split(',') if term.strip()]
        member_query = Q()
        for term in search_terms:
            member_query |= (
                Q(request_key__icontains=term) |
                Q(uname__icontains=term)
            )
        
        # 페이지네이션 적용
        offset = (page - 1) * limit
        members = Member.newspic_objects().filter(member_query).order_by('-level', 'uname')[offset:offset + limit + 1]
        
        # 다음 페이지 여부 확인
        has_more = len(members) > limit
        if has_more:
            members = members[:limit]
        
        member_data = []
        for member in members:
            member_data.append({
                'request_key': member.request_key,
                'uname': member.uname or '',
                'level': member.level,
                'uid': member.uid or ''
            })
        
        return JsonResponse({
            'members': member_data,
            'has_more': has_more,
            'page': page
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)